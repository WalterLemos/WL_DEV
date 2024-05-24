from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
from selenium.common.exceptions import NoSuchElementException
import pyautogui
from time import sleep
import openpyxl

def new_func():
    pyautogui.press('backspace',presses=20)

# entrar no site da -https://www.dividaativa.pge.sp.gov.br/sc/pages/consultas/consultarDebito.jsf
chrome_driver_path = r'C:\ProjetorPython\dev\chromedriver-win64\chromedriver.exe'
driver = webdriver.Chrome(executable_path=chrome_driver_path)
driver.get('https://www.dividaativa.pge.sp.gov.br/sc/pages/consultas/consultarDebito.jsf')

pyautogui.click(982,25)
sleep(0.5)
pyautogui.click(1225,29)
sleep(0.5)
#clicar nos 3 pontos
pyautogui.click(1892,80)
sleep(1)
#clicar em imprimir
pyautogui.click(1567,370)
sleep(0.5)
#clicar em destino
pyautogui.click(1489,203)
sleep(0.5)
#clicar em guardar como PDF
pyautogui.click(1477,262)
sleep(0.5)
#clicar em guardar
pyautogui.click(1438,913)
sleep(0.5)
#apagar nome
new_func()
#Escreve a CDA
pyautogui.write('exlcuir')
#clicar em salvar
pyautogui.click(994,638)
sleep(0.5)
#clicar em sim (caso tenha algo para substituir)
pyautogui.click(1027,541)

# Nome do arquivo Excel e nome da planilha
nome_arquivo_excel = r'C:\ProjetorPython\files\1505833-53.2018.8.26.0014.xlsx'
nome_planilha_excel = 'Débitos IPVA SP - Pan Arre'

# Carregar a planilha Excel
workbook = openpyxl.load_workbook(nome_arquivo_excel)
planilha = workbook[nome_planilha_excel]

# Começando da linha 4, coluna 3 (C)
start_row = 4
column_index = 3

# Descobrir o número total de linhas na planilha
total_rows = planilha.max_row

# Loop para consultar todas as linhas da coluna 3, começando da linha 4
row = start_row  # Inicialize a variável row
primeiro_registro = True  # Variável para controlar o primeiro registro

while row <= total_rows:
    Num_CDA = planilha.cell(row=row, column=column_index).value

# Verifique se Num_CDA está vazio
    if Num_CDA is None:
        print("Num_CDA está vazio. Saindo do loop.")
        break  # Sai do loop se Num_CDA estiver vazio

   # Digitar o CDA
    Campo_CDA = driver.find_element(By.XPATH, "//input[@id='consultaDebitoForm:decTxtTipoConsulta:cdaEtiqueta']")
    Campo_CDA.clear()  # Limpar o campo antes de inserir um novo valor
    Campo_CDA.send_keys(Num_CDA)
   
    # Se for o primeiro registro, espere 35 segundos
    if primeiro_registro:
        sleep(35)
        primeiro_registro = False  # Defina como False após o primeiro registro

    # Botão Consultar
    btn_Consultar = driver.find_element(By.XPATH, "//input[@name='consultaDebitoForm:j_id104']")
    btn_Consultar.click()
    sleep(1)

    # Verifique se a mensagem "Nenhum resultado com os critérios de consulta" está presente
    resultado_msg_element = driver.find_elements(By.XPATH, "//div[@class='rich-panel-body ']//p")
    resultado_msg = resultado_msg_element[0].text if resultado_msg_element else ""

   # Escreva a mensagem no Excel
    planilha.cell(row=row, column=column_index + 1, value=resultado_msg)

    if "Nenhum resultado com os critérios de consulta" in resultado_msg:
        CDA = str(Num_CDA)
        #Gerar PDF da Tela
        #clicar nos 3 pontos
        pyautogui.click(1892,80)
        sleep(1)
        #clicar em imprimir
        pyautogui.click(1567,370)
        sleep(0.5)
        #clicar em guardar
        pyautogui.click(1438,913)
        sleep(0.5)
        #apagar nome
        new_func()
        #Escreve a CDA
        pyautogui.write(CDA)
        #clicar em salvar
        pyautogui.click(994,638)
        sleep(0.5)
        #clicar em sim (caso tenha algo para substituir)
        pyautogui.click(1027,541)

        row += 1  # Vá para a próxima linha e continue o loop
        continue
    # Consultar IPVA
    href_IPVA = driver.find_element(By.XPATH, "//a[@id='consultaDebitoForm:dataTable:0:lnkConsultaDebito']")
    href_IPVA.click()
    sleep(1)

    # Consultar Registro
    href_Registro = driver.find_element(By.XPATH, "//a[@href='#']")
    href_Registro.click()
    sleep(1) 
    
   # Verifique se o elemento valor_honorarios_element existe
    try:
        valor_honorarios_element = driver.find_element(By.XPATH, "//tbody[@id='consultaDebitoForm:j_id1114:tb']//td[@id='consultaDebitoForm:j_id1114:4:j_id1122']")
        valor_honorarios = valor_honorarios_element.text

        valor_mora_multa_element = driver.find_element(By.XPATH,"//tbody[@id='consultaDebitoForm:j_id1114:tb']//td[@id='consultaDebitoForm:j_id1114:5:j_id1122']")
        valor_mora_multa = valor_mora_multa_element.text

    except NoSuchElementException:  
        valor_mora_multa_element = driver.find_element(By.XPATH,"//tbody[@id='consultaDebitoForm:j_id1114:tb']//td[@id='consultaDebitoForm:j_id1114:4:j_id1122']")
        valor_mora_multa = valor_mora_multa_element.text
        valor_honorarios = 0
      
    # Extrair informações
    num_registro_element = driver.find_element(By.XPATH, "//div[@id='consultaDebitoForm:j_id1008']")
    num_registro = num_registro_element.find_element(By.TAG_NAME, "span").text
    
    numero_processo_element = driver.find_element(By.XPATH, "//div[@id='consultaDebitoForm:j_id1024']")
    numero_processo = numero_processo_element.find_element(By.TAG_NAME, "span").text
    
    numero_processo_outros_element = driver.find_element(By.XPATH, "//div[@id='consultaDebitoForm:j_id1033']")
    numero_processo_outros = numero_processo_outros_element.find_element(By.TAG_NAME, "span").text

    data_inscricao_element = driver.find_element(By.XPATH, "//div[@id='consultaDebitoForm:j_id1016']")
    data_inscricao = data_inscricao_element.find_element(By.TAG_NAME, "span").text

    situacao_element = driver.find_element(By.XPATH, "//div[@id='consultaDebitoForm:j_id1090']")
    situacao = situacao_element.find_element(By.TAG_NAME, "span").text
    
    saldo_element = driver.find_element(By.XPATH, "//div[@id='consultaDebitoForm:j_id1106']")
    saldo= saldo_element.find_element(By.TAG_NAME, "span").text
    
    valor_principal_element = driver.find_element(By.XPATH,"//tbody[@id='consultaDebitoForm:j_id1114:tb']//tr[contains(@class,'rich-table-row')]//td[@id='consultaDebitoForm:j_id1114:0:j_id1122']")
    valor_principal = valor_principal_element.text

    valor_juros_element = driver.find_element(By.XPATH,"//tbody[@id='consultaDebitoForm:j_id1114:tb']//tr[contains(@class,'rich-table-row')]//td[@id='consultaDebitoForm:j_id1114:2:j_id1122']")
    valor_juros = valor_juros_element.text

    valor_multa_element = driver.find_element(By.XPATH,"//tbody[@id='consultaDebitoForm:j_id1114:tb']//td[@id='consultaDebitoForm:j_id1114:3:j_id1122']")
    valor_multa = valor_multa_element.text
        
    placa_element = driver.find_element(By.XPATH,"//tbody[@id='consultaDebitoForm:detalheDebitoIpvaDeclarado:tb']//tr[contains(@class,'rich-table-row')]//td[@id='consultaDebitoForm:detalheDebitoIpvaDeclarado:0:j_id1173']")
    placa = placa_element.text
    
    renavam_element = driver.find_element(By.XPATH,"//tbody[@id='consultaDebitoForm:detalheDebitoIpvaDeclarado:tb']//tr[contains(@class,'rich-table-row')]//td[@id='consultaDebitoForm:detalheDebitoIpvaDeclarado:0:j_id1175']")
    renavam = renavam_element.text
    
    chassi_element = driver.find_element(By.XPATH,"//tbody[@id='consultaDebitoForm:detalheDebitoIpvaDeclarado:tb']//tr[contains(@class,'rich-table-row')]//td[@id='consultaDebitoForm:detalheDebitoIpvaDeclarado:0:j_id1177']")
    chassi = chassi_element.text
   
    marca_modelo_element = driver.find_element(By.XPATH,"//tbody[@id='consultaDebitoForm:detalheDebitoIpvaDeclarado:tb']//tr[contains(@class,'rich-table-row')]//td[@id='consultaDebitoForm:detalheDebitoIpvaDeclarado:0:j_id1179']")
    marca_modelo = marca_modelo_element.text
    
    ano_fab_element = driver.find_element(By.XPATH,"//tbody[@id='consultaDebitoForm:detalheDebitoIpvaDeclarado:tb']//tr[contains(@class,'rich-table-row')]//td[@id='consultaDebitoForm:detalheDebitoIpvaDeclarado:0:j_id1181']")
    ano_fab = ano_fab_element.text
   
    ano_exercicio_element = driver.find_element(By.XPATH,"//tbody[@id='consultaDebitoForm:detalheDebitoIpvaDeclarado:tb']//tr[contains(@class,'rich-table-row')]//td[@id='consultaDebitoForm:detalheDebitoIpvaDeclarado:0:j_id1183']")
    ano_exercicio = ano_exercicio_element.text
    
    dt_parcelas_element = driver.find_element(By.XPATH,"//tbody[@id='consultaDebitoForm:detalheDebitoIpvaDeclarado:tb']//tr[contains(@class,'rich-table-row')]//td[@id='consultaDebitoForm:detalheDebitoIpvaDeclarado:0:j_id1185']")
    dt_parcelas = dt_parcelas_element.text
    
    # Encontre a próxima coluna disponível (vamos começar da coluna D)
    coluna_atual = column_index + 1


    # Escreva os valores nas colunas a partir da coluna D

    planilha.cell(row=row, column=coluna_atual, value=data_inscricao)
    coluna_atual += 1

    planilha.cell(row=row, column=coluna_atual, value=numero_processo)
    coluna_atual += 1

    planilha.cell(row=row, column=coluna_atual, value=numero_processo_outros)
    coluna_atual += 1

    planilha.cell(row=row, column=coluna_atual, value=situacao)
    coluna_atual += 1

    planilha.cell(row=row, column=coluna_atual, value=saldo)
    coluna_atual += 1

    planilha.cell(row=row, column=coluna_atual, value=valor_principal)
    coluna_atual += 1

    planilha.cell(row=row, column=coluna_atual, value=valor_juros)
    coluna_atual += 1

    planilha.cell(row=row, column=coluna_atual, value=valor_multa)
    coluna_atual += 1

    planilha.cell(row=row, column=coluna_atual, value=valor_honorarios)
    coluna_atual += 1

    planilha.cell(row=row, column=coluna_atual, value=valor_mora_multa)
    coluna_atual += 1

    planilha.cell(row=row, column=coluna_atual, value=placa)
    coluna_atual += 1

    planilha.cell(row=row, column=coluna_atual, value=renavam)
    coluna_atual += 1

    planilha.cell(row=row, column=coluna_atual, value=chassi)
    coluna_atual += 1

    planilha.cell(row=row, column=coluna_atual, value=marca_modelo)
    coluna_atual += 1

    planilha.cell(row=row, column=coluna_atual, value=ano_fab)
    coluna_atual += 1

    planilha.cell(row=row, column=coluna_atual, value=ano_exercicio)
    coluna_atual += 1

    planilha.cell(row=row, column=coluna_atual, value=dt_parcelas)
    
    # Salve o arquivo Excel
    workbook.save(nome_arquivo_excel)
    CDA = str(Num_CDA)
    #Gerar PDF da Tela
    #clicar nos 3 pontos
    pyautogui.click(1892,80)
    sleep(1)
    #clicar em imprimir
    pyautogui.click(1567,370)
    sleep(0.5)
    #clicar em guardar
    pyautogui.click(1438,913)
    sleep(0.5)
    #apagar nome
    new_func()
    #Escreve a CDA
    pyautogui.write(CDA)
    #clicar em salvar
    pyautogui.click(994,638)
    sleep(0.5)
    #clicar em sim (caso tenha algo para substituir)
    pyautogui.click(1027,541)


     # Botão Voltar
    btn_Voltar = driver.find_element(By.XPATH, "//input[@name='consultaDebitoForm:btnVoltarDetalheDebito']")
    btn_Voltar.click()
    sleep(2)

     # Botão Voltar
    btn_Voltar1 = driver.find_element(By.XPATH, "//input[@name='consultaDebitoForm:j_id264']")
    btn_Voltar1.click()
    sleep(2)
    row += 1

# Fechar o WebDriver
driver.quit()

