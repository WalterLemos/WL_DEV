from selenium import webdriver
import os
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pyautogui
from time import sleep
import openpyxl

def new_func():
    pyautogui.press('backspace',presses=20)

def printlongo():
  pyautogui.keyDown('ctrl')
  pyautogui.press('p')
  pyautogui.keyUp('ctrl')
  sleep(1)
  #clicar em destino
  pyautogui.click(1461,229)
  sleep(0.5)
  #clicar em guardar como PDF
  pyautogui.click(1376,271)
  sleep(0.5)
  #clicar em guardar
  pyautogui.click(1434,908)
  sleep(0.5)
  #apagar nome
  new_func()
  #Escreve a CDA
  pyautogui.write('excluir')
  #clicar em salvar
  pyautogui.click(696,741)
  sleep(0.5)
  #clicar em sim (caso tenha algo para substituir)
  pyautogui.click(1020,527)

def printcurto():
  pyautogui.keyDown('ctrl')
  pyautogui.press('p')
  pyautogui.keyUp('ctrl')
  sleep(1)
  #clicar em guardar
  pyautogui.click(1434,908)
  sleep(0.5)
  #apagar nome
  new_func()
  #Escreve a CDA
  pyautogui.write(CDA)
  #clicar em salvar
  pyautogui.click(696,741)
  sleep(0.5)
  #clicar em sim (caso tenha algo para substituir)
  pyautogui.click(1020,527)

## Configurando o caminho do executável como variável de ambiente
chrome_driver_path = r'C:\Users\suporte\Desktop\PSTs j&f C:\Users\suporte\Desktop\PSTs j&f jbs\chromedriver-win64\chromedriver.exe'
os.environ["webdriver.chrome.driver"] = chrome_driver_path

# Inicializando o driver do Chrome
driver = webdriver.Chrome()
driver.get('https://www.dividaativa.pge.sp.gov.br/sc/pages/consultas/consultarDebito.jsf')

primeiro_registro2 = True  # Variável para controlar o primeiro registro

pyautogui.click(1235,43) #Clicar para expandir o chrome
sleep(5)
pyautogui.click(1883,153) #Clicar para fechar a notificação de barra de tarefas

if primeiro_registro2:
   printlongo()
   primeiro_registro2 = False  # Defina como False após o primeiro registro 
   sleep(1)

# Nome do arquivo Excel e nome da planilha
nome_arquivo_excel = r'C:\Users\suporte\Desktop\Rõbo MDU\bichara-robos\Demanda 0159.xlsx'
nome_planilha_excel = 'Débitos IPVA'

# Carregar a planilha Excel
workbook = openpyxl.load_workbook(nome_arquivo_excel)
planilha = workbook[nome_planilha_excel]

# Começando da linha 4, coluna 3 (C)
start_row = 660
column_index = 3

# Descobrir o número total de linhas na planilha
total_rows = planilha.max_row

# Loop para consultar todas as linhas da coluna 3, começando da linha 4
row = start_row  # Inicialize a variável row
primeiro_registro = True  # Variável para controlar o primeiro registro

while row <= total_rows:
    Num_CDA = planilha.cell(row=row, column=column_index).value

# Verifique se Num_CDA está vazio
    if Num_CDA is None:
        print("Num_CDA está vazio. Saindo do loop.")
        break  # Sai do loop se Num_CDA estiver vazio

   # Digitar o CDA
    Campo_CDA = driver.find_element(By.XPATH, "//input[@id='consultaDebitoForm:decTxtTipoConsulta:cdaEtiqueta']")
    Campo_CDA.clear()  # Limpar o campo antes de inserir um novo valor
    Campo_CDA.send_keys(Num_CDA)
   
    # Se for o primeiro registro, espere 35 segundos
    if primeiro_registro:
        sleep(35)
        primeiro_registro = False  # Defina como False após o primeiro registro

    # Botão Consultar
    btn_Consultar = driver.find_element(By.XPATH, "//input[@name='consultaDebitoForm:j_id114']")
    btn_Consultar.click()
    sleep(1)

    # Verifique se a mensagem "Nenhum resultado com os critérios de consulta" está presente
    resultado_msg_element = driver.find_elements(By.XPATH, "//div[@class='rich-panel-body ']//p")
    resultado_msg = resultado_msg_element[0].text if resultado_msg_element else ""

   # Escreva a mensagem no Excel
    planilha.cell(row=row, column=column_index + 1, value=resultado_msg)

    if "Nenhum resultado com os critérios de consulta" in resultado_msg:
        CDA = str(Num_CDA)
        printcurto()
        row += 1  # Vá para a próxima linha e continue o loop
        continue
    # Consultar IPVA
    href_IPVA = driver.find_element(By.XPATH, "//a[@id='consultaDebitoForm:dataTable:0:lnkConsultaDebito']")
    href_IPVA.click()
    sleep(1)

    # Consultar Registro
    href_Registro = driver.find_element(By.XPATH, "//a[@href='#']")
    href_Registro.click()
    sleep(1) 
    
   # Verifique se o elemento valor_honorarios_element exis
    try:
      valor_honorarios_element = driver.find_element(By.XPATH, "//tbody[@id='consultaDebitoForm:j_id1128:tb']//td[@id='consultaDebitoForm:j_id1128:4:j_id1136']")
      valor_honorarios = valor_honorarios_element.text
    except NoSuchElementException:
       valor_honorarios = 0
    try:
      valor_mora_multa_element = driver.find_element(By.XPATH,"//tbody[@id='consultaDebitoForm:j_id1128:tb']//td[@id='consultaDebitoForm:j_id1128:5:j_id1136']")
      valor_mora_multa = valor_mora_multa_element.text
    except NoSuchElementException:
      valor_mora_multa = 0

       # Extrair informações
    num_registro_element = driver.find_element(By.XPATH, "//div[@id='consultaDebitoForm:j_id1022']")
    num_registro = num_registro_element.find_element(By.TAG_NAME, "span").text
    
    numero_processo_element = driver.find_element(By.XPATH, "//div[@id='consultaDebitoForm:j_id1038']")
    numero_processo = numero_processo_element.find_element(By.TAG_NAME, "span").text
    
    numero_processo_outros_element = driver.find_element(By.XPATH, "//div[@id='consultaDebitoForm:j_id1047']")
    numero_processo_outros = numero_processo_outros_element.find_element(By.TAG_NAME, "span").text

    data_inscricao_element = driver.find_element(By.XPATH, "//div[@id='consultaDebitoForm:j_id1030']")
    data_inscricao = data_inscricao_element.find_element(By.TAG_NAME, "span").text

    situacao_element = driver.find_element(By.XPATH, "//div[@id='consultaDebitoForm:j_id1104']")
    situacao = situacao_element.find_element(By.TAG_NAME, "span").text
    
    saldo_element = driver.find_element(By.XPATH, "//div[@id='consultaDebitoForm:j_id1120']")
    saldo= saldo_element.find_element(By.TAG_NAME, "span").text
    
    valor_principal_element = driver.find_element(By.XPATH,"//tbody[@id='consultaDebitoForm:j_id1128:tb']//td[@id='consultaDebitoForm:j_id1128:0:j_id1136']")
    valor_principal = valor_principal_element.text

    valor_juros_element = driver.find_element(By.XPATH,"//tbody[@id='consultaDebitoForm:j_id1128:tb']//td[@id='consultaDebitoForm:j_id1128:2:j_id1136']")
    valor_juros = valor_juros_element.text

    valor_multa_element = driver.find_element(By.XPATH,"//tbody[@id='consultaDebitoForm:j_id1128:tb']//td[@id='consultaDebitoForm:j_id1128:3:j_id1136']")
    valor_multa = valor_multa_element.text
        
    placa_element = driver.find_element(By.XPATH,"//tbody[@id='consultaDebitoForm:detalheDebitoIpvaDeclarado:tb']//td[@id='consultaDebitoForm:detalheDebitoIpvaDeclarado:0:j_id1187']")
    placa = placa_element.text
    
    renavam_element = driver.find_element(By.XPATH,"//tbody[@id='consultaDebitoForm:detalheDebitoIpvaDeclarado:tb']//td[@id='consultaDebitoForm:detalheDebitoIpvaDeclarado:0:j_id1189']")
    renavam = renavam_element.text
    
    chassi_element = driver.find_element(By.XPATH,"//tbody[@id='consultaDebitoForm:detalheDebitoIpvaDeclarado:tb']//td[@id='consultaDebitoForm:detalheDebitoIpvaDeclarado:0:j_id1191']")
    chassi = chassi_element.text
   
    marca_modelo_element = driver.find_element(By.XPATH,"//tbody[@id='consultaDebitoForm:detalheDebitoIpvaDeclarado:tb']//td[@id='consultaDebitoForm:detalheDebitoIpvaDeclarado:0:j_id1193']")
    marca_modelo = marca_modelo_element.text
    
    ano_fab_element = driver.find_element(By.XPATH,"//tbody[@id='consultaDebitoForm:detalheDebitoIpvaDeclarado:tb']//td[@id='consultaDebitoForm:detalheDebitoIpvaDeclarado:0:j_id1195']")
    ano_fab = ano_fab_element.text
   
    ano_exercicio_element = driver.find_element(By.XPATH,"//tbody[@id='consultaDebitoForm:detalheDebitoIpvaDeclarado:tb']//td[@id='consultaDebitoForm:detalheDebitoIpvaDeclarado:0:j_id1197']")
    ano_exercicio = ano_exercicio_element.text
    
    dt_parcelas_element = driver.find_element(By.XPATH,"//tbody[@id='consultaDebitoForm:detalheDebitoIpvaDeclarado:tb']//td[@id='consultaDebitoForm:detalheDebitoIpvaDeclarado:0:j_id1199']")
    dt_parcelas = dt_parcelas_element.text
  
    
    # Encontre a próxima coluna disponível (vamos começar da coluna D)
    coluna_atual = column_index + 1

    # Escreva os valores nas colunas a partir da coluna D

    planilha.cell(row=row, column=coluna_atual, value=data_inscricao)
    coluna_atual += 1

    planilha.cell(row=row, column=coluna_atual, value=numero_processo)
    coluna_atual += 1

    planilha.cell(row=row, column=coluna_atual, value=numero_processo_outros)
    coluna_atual += 1

    planilha.cell(row=row, column=coluna_atual, value=situacao)
    coluna_atual += 1

    planilha.cell(row=row, column=coluna_atual, value=saldo)
    coluna_atual += 1

    planilha.cell(row=row, column=coluna_atual, value=valor_principal)
    coluna_atual += 1

    planilha.cell(row=row, column=coluna_atual, value=valor_juros)
    coluna_atual += 1

    planilha.cell(row=row, column=coluna_atual, value=valor_multa)
    coluna_atual += 1

    planilha.cell(row=row, column=coluna_atual, value=valor_honorarios)
    coluna_atual += 1

    planilha.cell(row=row, column=coluna_atual, value=valor_mora_multa)
    coluna_atual += 1

    planilha.cell(row=row, column=coluna_atual, value=placa)
    coluna_atual += 1

    planilha.cell(row=row, column=coluna_atual, value=renavam)
    coluna_atual += 1

    planilha.cell(row=row, column=coluna_atual, value=chassi)
    coluna_atual += 1

    planilha.cell(row=row, column=coluna_atual, value=marca_modelo)
    coluna_atual += 1

    planilha.cell(row=row, column=coluna_atual, value=ano_fab)
    coluna_atual += 1

    planilha.cell(row=row, column=coluna_atual, value=ano_exercicio)
    coluna_atual += 1

    planilha.cell(row=row, column=coluna_atual, value=dt_parcelas)
    
    # Salve o arquivo Excel
    workbook.save(nome_arquivo_excel)
    CDA = str(Num_CDA)
    printcurto()

    # Botão Voltar
    wait = WebDriverWait(driver, 15)
    btn_Voltar = wait.until(EC.presence_of_element_located((By.XPATH, "//input[@id='consultaDebitoForm:btnVoltarDetalheDebito']")))
    #driver.find_element(By.XPATH, "//input[@name='consultaDebitoForm:btnVoltarDetalheDebito']")
    btn_Voltar.click()
    sleep(2)

     # Botão Voltar
    wait = WebDriverWait(driver, 15)
    btn_Voltar1 = wait.until(EC.presence_of_element_located((By.XPATH, "//input[@name='consultaDebitoForm:j_id278']")))
      #driver.find_element(By.XPATH, "//input[@name='consultaDebitoForm:j_id264']")
    btn_Voltar1.click()
    sleep(2)
    row += 1

# Fechar o WebDriver
driver.quit()

