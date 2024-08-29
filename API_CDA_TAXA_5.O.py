import os
import base64
import tkinter as tk
from tkinter import PhotoImage
from tkinter import filedialog, messagebox
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.select import Select
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl
from time import sleep
from time import sleep
import re
from selenium.common.exceptions import NoSuchElementException
from twocaptcha import TwoCaptcha
from anticaptchaofficial.recaptchav2proxyless import *
from Chave import *

def configurar_chrome_options():
    chrome_options = Options()
    chrome_options.add_argument('--kiosk-printing')  # Ativa a impressão automática sem diálogo
    return chrome_options

def gerar_pdf_dinamico(driver, caminho_diretorio, nome_pdf):
    # Define o caminho completo do arquivo PDF usando o nome_pdf fornecido
    caminho_arquivo_pdf = os.path.join(caminho_diretorio, f"{nome_pdf}.pdf")

    # Usa a API DevTools para gerar o PDF da página atual
    result = driver.execute_cdp_cmd("Page.printToPDF", {
        "landscape": False,                 # Define orientação do PDF como retrato
        "displayHeaderFooter": False,       # Não exibe cabeçalho e rodapé
        "printBackground": True,            # Inclui o fundo da página no PDF
        "preferCSSPageSize": True           # Usa o tamanho da página definido por CSS
    })

    # Decodifica o PDF gerado (base64) e salva em um arquivo
    with open(caminho_arquivo_pdf, "wb") as file:
        file.write(base64.b64decode(result['data']))

    print(f"PDF gerado e salvo em: {caminho_arquivo_pdf}")

def start_process(excel_path, sheet_name, start_row, output_dir):
    # Configurando o Chrome com as opções de impressão
    chrome_options = configurar_chrome_options()
    driver_service = Service((ChromeDriverManager().install()))
    driver = webdriver.Chrome(service=driver_service, options=chrome_options)
    driver.maximize_window()

    link = 'https://www.dividaativa.pge.sp.gov.br/sc/pages/consultas/consultarDebito.jsf'
    driver.get(link)
    sleep(2)

    primeiro_registro2 = True  # Variável para controlar o primeiro registro

    # Carregar a planilha Excel
    workbook = openpyxl.load_workbook(excel_path)
    planilha = workbook[sheet_name]

    column_index = 3
    total_rows = planilha.max_row
    row = 4
    primeiro_registro = True

    while row <= total_rows:
    
        # Verificação se existe um OK na linha atual coluna 21 (coluna de Verificação)
        while row <= total_rows:
            valor_coluna_21 = planilha.cell(row=row, column=21).value
            
            #Se exister o OK ele passa para próxima linha da tabela até achar uma célula vazia para continuar
            if valor_coluna_21 == "OK":
                row += 1
            else:
                break

        # Se a linha ultrapassou o total de linhas, interrompe o loop
        if row > total_rows:
            break
        
        Num_CDA = planilha.cell(row=row, column=column_index).value

        if Num_CDA is None:
            print("Num_CDA está vazio. Saindo do loop.")
            break
        
        Campo_CDA = driver.find_element(By.XPATH, "//input[@id='consultaDebitoForm:decTxtTipoConsulta:cdaEtiqueta']")
        Campo_CDA.clear()
        Campo_CDA.send_keys(Num_CDA)

        # Se for o primeiro registro, espere 35 segundos
        if primeiro_registro:
            chave_captcha = driver.find_element(By.CLASS_NAME, 'g-recaptcha').get_attribute('data-sitekey')

            solver = recaptchaV2Proxyless()
            solver.set_verbose(1)
            solver.set_key(chave_api)
            solver.set_website_url(link)
            solver.set_website_key(chave_captcha)

            resposta = solver.solve_and_return_solution()

            if resposta != 0:
                #print(resposta)
                # preencher o campo do token do captcha
                driver.execute_script(f"document.getElementById('g-recaptcha-response').innerHTML = '{resposta}'")
                driver.find_element(By.XPATH, "//*[@id='consultaDebitoForm:j_id78_body']/div[2]/input[2]").click()
                primeiro_registro = False  # Defina como False após o primeiro registro
            else:
                print(solver.err_string)

        btn_Consultar = driver.find_element(By.XPATH, "//*[@id='consultaDebitoForm:j_id78_body']/div[2]/input[2]")
        btn_Consultar.click()
        sleep(1)

        resultado_msg_element = driver.find_elements(By.XPATH, "//div[@class='rich-panel-body ']//p")
        resultado_msg = resultado_msg_element[0].text if resultado_msg_element else ""

        if "Nenhum resultado com os critérios de consulta" in resultado_msg:
            planilha.cell(row=row, column=column_index + 1, value="Nenhum resultado com os critérios de consulta")
            planilha.cell(row=row, column=21, value="OK") #Escreve "OK" na linha atual na coluna 21 (Coluna de cerificação)
            workbook.save(excel_path)
            
            # Aqui você chama a função para gerar e salvar o PDF com o nome da CDA
            gerar_pdf_dinamico(driver, output_dir, Num_CDA)
            
            row += 1
            continue

            # Consultar IPVA
        href_IPVA = driver.find_element(By.XPATH, "//a[@id='consultaDebitoForm:dataTable:0:lnkConsultaDebito']")
        href_IPVA.click()
        sleep(1)

        # Consultar Registro
        href_Registro = driver.find_element(By.XPATH, "//a[@href='#']")
        href_Registro.click()
        sleep(1) 
        try:
            correcao_monetaria_element = driver.find_element(By.XPATH, "//td[@id='consultaDebitoForm:j_id1128:1:j_id1136']")
            corrcao_monetaria = correcao_monetaria_element.text
        except NoSuchElementException:
            corrcao_monetaria = 0

        # Extrair informações
        num_registro_element = driver.find_element(By.XPATH, "//*[@id='consultaDebitoForm:j_id1028']")
        num_registro = num_registro_element.find_element(By.TAG_NAME, "span").text
        
        numero_processo_element = driver.find_element(By.XPATH, "//*[@id='consultaDebitoForm:j_id1044']")
        numero_processo = numero_processo_element.find_element(By.TAG_NAME, "span").text
        
        numero_processo_outros_element = driver.find_element(By.XPATH, "//*[@id='consultaDebitoForm:j_id1053']")
        numero_processo_outros = numero_processo_outros_element.find_element(By.TAG_NAME, "span").text

        numero_processo_judicial_element = driver.find_element(By.XPATH, "//*[@id='consultaDebitoForm:j_id1062']")
        numero_processo_judicial = numero_processo_judicial_element.find_element(By.TAG_NAME, "span").text

        parte_reu_element = driver.find_element(By.XPATH, "//*[@id='consultaDebitoForm:j_id1078']")
        parte_reu = parte_reu_element.find_element(By.TAG_NAME, "span").text

        parte_autor_element = driver.find_element(By.XPATH, "//*[@id='consultaDebitoForm:j_id1070']")
        parte_antor = parte_autor_element.find_element(By.TAG_NAME, "span").text

        vara_element = driver.find_element(By.XPATH, "//*[@id='consultaDebitoForm:j_id1086']")
        vara = vara_element.find_element(By.TAG_NAME, "span").text

        comarca_foro_element = driver.find_element(By.XPATH, "//*[@id='consultaDebitoForm:j_id1094']")
        comarca_foro = comarca_foro_element.find_element(By.TAG_NAME, "span").text

        data_inscricao_element = driver.find_element(By.XPATH, "//*[@id='consultaDebitoForm:j_id1036']")
        data_inscricao = data_inscricao_element.find_element(By.TAG_NAME, "span").text

        situacao_element = driver.find_element(By.XPATH, "//*[@id='consultaDebitoForm:j_id1110']")
        situacao = situacao_element.find_element(By.TAG_NAME, "span").text
        
        saldo_element = driver.find_element(By.XPATH, "//*[@id='consultaDebitoForm:j_id1126']")
        saldo= saldo_element.find_element(By.TAG_NAME, "span").text
        
        valor_principal_element = driver.find_element(By.XPATH,"//*[@id='consultaDebitoForm:j_id1134:0:j_id1142']")
        valor_principal = valor_principal_element.text

        valor_juros_element = driver.find_element(By.XPATH,"//*[@id='consultaDebitoForm:j_id1134:2:j_id1142']")
        valor_juros = valor_juros_element.text

        valor_honorarios_element = driver.find_element(By.XPATH, "//*[@id='consultaDebitoForm:j_id1134:3:j_id1142']")
        valor_honorarios = valor_honorarios_element.text

        data_element = driver.find_element(By.XPATH,"//*[@id='consultaDebitoForm:j_id1154:0:j_id1169']")
        data = data_element.text
        
        valor_element = driver.find_element(By.XPATH,"//*[@id='consultaDebitoForm:j_id1154:0:j_id1171']")
        valor= valor_element.text
        
        data_inicio_juros_element = driver.find_element(By.XPATH,"//*[@id='consultaDebitoForm:j_id1154:0:j_id1174']")
        data_inicio_juros = data_inicio_juros_element.text
    
        data_inicio_correcao_element = driver.find_element(By.XPATH,"//*[@id='consultaDebitoForm:j_id1154:0:j_id1176']")
        data_inicio_correcao = data_inicio_correcao_element.text
        
        # Encontre a próxima coluna disponível (vamos começar da coluna D)
        coluna_atual = column_index + 1

        # Escreva os valores nas colunas a partir da coluna D

        planilha.cell(row=row, column=coluna_atual, value=data_inscricao)
        coluna_atual += 1

        planilha.cell(row=row, column=coluna_atual, value=numero_processo)
        coluna_atual += 1

        planilha.cell(row=row, column=coluna_atual, value=numero_processo_outros)
        coluna_atual += 1

        planilha.cell(row=row, column=coluna_atual, value=numero_processo_judicial)
        coluna_atual += 1

        planilha.cell(row=row, column=coluna_atual, value=parte_reu)
        coluna_atual += 1

        planilha.cell(row=row, column=coluna_atual, value=parte_antor)
        coluna_atual += 1

        planilha.cell(row=row, column=coluna_atual, value=comarca_foro)
        coluna_atual += 1

        planilha.cell(row=row, column=coluna_atual, value=vara)
        coluna_atual += 1

        planilha.cell(row=row, column=coluna_atual, value=situacao)
        coluna_atual += 1

        planilha.cell(row=row, column=coluna_atual, value=saldo)
        coluna_atual += 1

        planilha.cell(row=row, column=coluna_atual, value=valor_principal)
        coluna_atual += 1

        planilha.cell(row=row, column=coluna_atual, value=corrcao_monetaria)
        coluna_atual += 1

        planilha.cell(row=row, column=coluna_atual, value=valor_juros)
        coluna_atual += 1

        planilha.cell(row=row, column=coluna_atual, value=valor_honorarios)
        coluna_atual += 1

        planilha.cell(row=row, column=coluna_atual, value=data)
        coluna_atual += 1

        planilha.cell(row=row, column=coluna_atual, value=valor)
        coluna_atual += 1

        planilha.cell(row=row, column=coluna_atual, value=data_inicio_juros)
        coluna_atual += 1

        planilha.cell(row=row, column=coluna_atual, value=data_inicio_correcao)
        coluna_atual += 1
        workbook.save(excel_path)
        CDA = str(Num_CDA)
    
        # Aqui você chama a função para gerar e salvar o PDF com o nome da CDA
        gerar_pdf_dinamico(driver, output_dir, Num_CDA)

        # Botão Voltar
        wait = WebDriverWait(driver, 15)
        btn_Voltar = wait.until(EC.presence_of_element_located((By.XPATH, "//*[@id='consultaDebitoForm:btnVoltarDetalheDebito']")))
        #driver.find_element(By.XPATH, "//input[@name='consultaDebitoForm:btnVoltarDetalheDebito']")
        btn_Voltar.click()
        sleep(2)

        # Botão Voltar
        wait = WebDriverWait(driver, 15)
        btn_Voltar1 = wait.until(EC.presence_of_element_located((By.XPATH, "//input[@name='consultaDebitoForm:j_id284']")))
        #driver.find_element(By.XPATH, "//input[@name='consultaDebitoForm:j_id284']")
        btn_Voltar1.click()
        sleep(2)
        row += 1

    driver.quit()

def browse_file(entry):
    filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
    entry.delete(0, tk.END)
    entry.insert(0, filename)

def browse_directory(entry):
    directory = filedialog.askdirectory()
    entry.delete(0, tk.END)
    entry.insert(0, directory)

def main():
    root = tk.Tk()
    root.title("Automação PGE SP")

    tk.Label(root, text="Arquivo Excel:").grid(row=0, column=0, sticky=tk.W, padx=10, pady=5)
    excel_entry = tk.Entry(root, width=50)
    excel_entry.grid(row=0, column=1, padx=10, pady=5)
    tk.Button(root, text="Procurar", command=lambda: browse_file(excel_entry)).grid(row=0, column=2, padx=(10, 150), pady=5, sticky=tk.E)

    tk.Label(root, text="Nome da Planilha:").grid(row=1, column=0, sticky=tk.W, padx=10, pady=5)
    sheet_entry = tk.Entry(root, width=50)
    sheet_entry.grid(row=1, column=1, padx=10, pady=5)

    # Entry para o texto copiável
    copyable_text = tk.Entry(root, width=30)
    copyable_text.insert(0, "Débitos IPVA SP - Pan Arre")
    copyable_text.config(state='readonly')  # Torna o campo somente leitura, mas ainda copiável
    copyable_text.grid(row=1, column=2, padx=10, pady=5)

    tk.Label(root, text="Diretório de Saída dos PDFs:").grid(row=3, column=0, sticky=tk.W, padx=10, pady=5)
    output_dir_entry = tk.Entry(root, width=50)
    output_dir_entry.grid(row=3, column=1, padx=10, pady=5)
    tk.Button(root, text="Procurar", command=lambda: browse_directory(output_dir_entry)).grid(row=3, column=2, padx=(10, 150), pady=5, sticky=tk.E)

    def on_start():
        excel_path = excel_entry.get()
        sheet_name = sheet_entry.get()
        try:
            start_row = 3
        except ValueError:
            messagebox.showerror("Erro", "A linha de início deve ser um número inteiro.")
            return
        output_dir = output_dir_entry.get()

        if not os.path.exists(excel_path):
            messagebox.showerror("Erro", "O arquivo Excel não existe.")
            return
        if not os.path.exists(output_dir):
            messagebox.showerror("Erro", "O diretório de saída não existe.")
            return

        start_process(excel_path, sheet_name, start_row, output_dir)
        messagebox.showinfo("Concluído", "Processo finalizado com sucesso.")

    tk.Button(root, text="Iniciar", command=on_start).grid(row=4, column=1, pady=10)

    root.mainloop()

if __name__ == "__main__":
    main()
                