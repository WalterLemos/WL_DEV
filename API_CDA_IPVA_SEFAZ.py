from selenium import webdriver
import os
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
import pyautogui
from time import sleep
import openpyxl
from tqdm import tqdm  # Importar a biblioteca tqdm

def new_func():
    pyautogui.press('backspace', presses=20)

def Imprimir(Renavam):
    
    pyautogui.keyDown('ctrl')
    pyautogui.press('p')
    pyautogui.keyUp('ctrl')
    sleep(1)
    #clicar em destino
    pyautogui.click(1505,269)
    sleep(0.5)
    #clicar em guardar como PDF
    pyautogui.click(1470,322)
    sleep(0.5)
    #clicar em guardar
    pyautogui.click(1484,918)
    sleep(0.5)
    #apagar nome
    new_func()
    #Escreve a CDA
    pyautogui.write(Renavam)
    #clicar em salvar
    pyautogui.click(711,675)
    sleep(0.5)
    #clicar em sim (caso tenha algo para substituir)
    pyautogui.click(1020,537)


def consulta_ipva():
    # Configurando o caminho do executável como variável de ambiente
    chrome_driver_path = r'C:\ProjetorPython\dev\chromedriver-win64\chromedriver.exe'
    os.environ["webdriver.chrome.driver"] = chrome_driver_path

    # Inicializando o driver do Chrome
    driver = webdriver.Chrome()
    driver.get('https://www.ipva.fazenda.sp.gov.br/IPVANET_Consulta/Consulta.aspx')

    pyautogui.click(1483,48)
   
    # Nome do arquivo Excel e nome da planilha
    nome_arquivo_excel = r'C:\Users\walter.oliveira\Documents\ProjetosPython\dev\Bichara_Dev\repository\Templete_Renavam.xlsx'
    nome_planilha_excel = 'Débitos Renavam'

    # Carregar a planilha Excel
    workbook = openpyxl.load_workbook(nome_arquivo_excel)
    planilha = workbook[nome_planilha_excel]

    # Começando da linha 4, coluna 3 (C)
    start_row = 4
    column_index = 3
    column_index2 = 4

    # Descobrir o número total de linhas na planilha
    total_rows = planilha.max_row

    # Loop para consultar todas as linhas da coluna 3, começando da linha 4
    row = start_row  # Inicialize a variável row

    # Inicializando a barra de progresso
    with tqdm(total=total_rows - start_row + 1, desc="Processando") as pbar:
        while row <= total_rows:
            Num_Renavam = planilha.cell(row=row, column=column_index).value
            Num_Placa = planilha.cell(row=row, column=column_index2).value

            # Verificar se Num_CDA está vazio
            if Num_Renavam is None:
                print("Num_Renavam está vazio. Saindo do loop.")
                break  # Sai do loop se Num_CDA estiver vazio

            try:
                # Digitar o Renavam
                Campo_Renavam = driver.find_element(By.XPATH, "//input[@id='conteudoPaginaPlaceHolder_txtRenavam']")
                Campo_Renavam.clear()  # Limpar o campo antes de inserir um novo valor
                Campo_Renavam.send_keys(Num_Renavam)

                # Digitar a Placa
                Campo_Placa = driver.find_element(By.XPATH, "//input[@id='conteudoPaginaPlaceHolder_txtPlaca']")
                Campo_Placa.clear()  # Limpar o campo antes de inserir um novo valor
                Campo_Placa.send_keys(Num_Placa)

                # Botão Consultar
                btn_Consultar = driver.find_element(By.XPATH, "//input[@id='conteudoPaginaPlaceHolder_btn_Consultar']")
                btn_Consultar.click()
                # Esperar a página carregar
                sleep(20)

                # Extrair informações
                valor_IPVA_Atual_element = driver.find_element(By.XPATH, "//span[@id='conteudoPaginaPlaceHolder_txtValoraPagar']")
                valor_IPVA_Atual = valor_IPVA_Atual_element.text

                ipva_debitos_element = driver.find_element(By.XPATH, "//span[@id='conteudoPaginaPlaceHolder_txtAnoPend01']")
                ipva_debitos = ipva_debitos_element.text

                ipva_divida_element = driver.find_element(By.XPATH, "//span[@id='conteudoPaginaPlaceHolder_txtExisteDividaAtiva']")
                ipva_divida_ativa = ipva_divida_element.text

                dpvat_element = driver.find_element(By.XPATH, "//span[@id='conteudoPaginaPlaceHolder_txtAnoDpvat01']") 
                dpvat = dpvat_element.text 

                multas_element = driver.find_element(By.XPATH, "//span[@id='conteudoPaginaPlaceHolder_txtResumoMultas']") 
                multas = multas_element.text

                license_2024_element = driver.find_element(By.XPATH, "//*[@id='conteudoPaginaPlaceHolder_tbTaxasDetalhe']/tbody/tr[6]/td[5]/span") 
                license_2024 = license_2024_element.text  

                license_2023_element = driver.find_element(By.XPATH, "//*[@id='conteudoPaginaPlaceHolder_tbTaxasDetalhe']/tbody/tr[11]/td[5]/span") 
                license_2023 = license_2023_element.text   

                license_2022_element = driver.find_element(By.XPATH, "//*[@id='conteudoPaginaPlaceHolder_tbTaxasDetalhe']/tbody/tr[10]/td[5]/span") 
                license_2022 = license_2022_element.text   

                license_2021_element = driver.find_element(By.XPATH, "//*[@id='conteudoPaginaPlaceHolder_tbTaxasDetalhe']/tbody/tr[9]/td[5]/span") 
                license_2021 = license_2021_element.text   

                license_2020_element = driver.find_element(By.XPATH, "//*[@id='conteudoPaginaPlaceHolder_tbTaxasDetalhe']/tbody/tr[8]/td[5]/span") 
                license_2020 = license_2020_element.text   

                license_2019_element = driver.find_element(By.XPATH, "//*[@id='conteudoPaginaPlaceHolder_tbTaxasDetalhe']/tbody/tr[7]/td[5]/span") 
                license_2019 = license_2019_element.text

                valor_Total_element = driver.find_element(By.XPATH, "//span[@id='conteudoPaginaPlaceHolder_txtValorTotalDebitos']")
                valor_Total = valor_Total_element.text    
                
                # Encontre a próxima coluna disponível (vamos começar da coluna E)
                coluna_atual = column_index + 2

                # Escreva os valores nas colunas a partir da coluna E
                planilha.cell(row=row, column=coluna_atual, value=valor_IPVA_Atual)
                coluna_atual += 1

                planilha.cell(row=row, column=coluna_atual, value=ipva_debitos)
                coluna_atual += 1

                planilha.cell(row=row, column=coluna_atual, value=ipva_divida_ativa)
                coluna_atual += 1

                planilha.cell(row=row, column=coluna_atual, value=dpvat)
                coluna_atual += 1

                planilha.cell(row=row, column=coluna_atual, value=multas)
                coluna_atual += 1

                planilha.cell(row=row, column=coluna_atual, value=license_2024)
                coluna_atual += 1 

                planilha.cell(row=row, column=coluna_atual, value=license_2023)
                coluna_atual += 1

                planilha.cell(row=row, column=coluna_atual, value=license_2022)
                coluna_atual += 1

                planilha.cell(row=row, column=coluna_atual, value=license_2021)
                coluna_atual += 1

                planilha.cell(row=row, column=coluna_atual, value=license_2020)
                coluna_atual += 1

                planilha.cell(row=row, column=coluna_atual, value=license_2019)
                coluna_atual += 1

                planilha.cell(row=row, column=coluna_atual, value=valor_Total)
                coluna_atual += 1
             
                Imprimir(Num_Renavam)
            
            except NoSuchElementException as e:
                print(f"Elemento não encontrado: {e}")

            row += 1  # Avança para a próxima linha
            pbar.update(1)  # Atualiza a barra de progresso

    # Salvar o arquivo Excel após todas as alterações
    workbook.save(nome_arquivo_excel)
    workbook.close()

    # Fechar o WebDriver
    driver.quit()

# Chamada da função principal
consulta_ipva()
