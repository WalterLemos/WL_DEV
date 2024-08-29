from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pyautogui
from time import sleep
import openpyxl
from anticaptchaofficial.recaptchav2proxyless import recaptchaV2Proxyless
from twocaptcha import TwoCaptcha
from Chave import *

def new_func():
    pyautogui.press('backspace', presses=20)

def salvar_como_pdf(CDA):
    pyautogui.hotkey('ctrl', 'p')
    sleep(2)
    pyautogui.click(1519, 259)
    sleep(1)
    pyautogui.click(1477, 330)
    sleep(1)
    pyautogui.click(1493, 918)
    sleep(1)
    new_func()
    pyautogui.write(f'relatorio_{CDA}.pdf')
    pyautogui.click(716, 670)
    sleep(1)
    pyautogui.click(1033, 527)

# Inicializando o driver do Chrome
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
link = 'https://www.dividaativa.pge.sp.gov.br/sc/pages/consultas/consultarDebito.jsf'
driver.get(link)

pyautogui.click(1480, 39)
sleep(1)
pyautogui.click(1868, 184)
sleep(1)

# Nome do arquivo Excel e nome da planilha
nome_arquivo_excel = r'C:\Users\walter.oliveira\Documents\ProjetosPython\dev\Bichara_Dev\repository\Templete_Taxa_Judiciaria.xlsx'
nome_planilha_excel = 'Débitos Taxa Judiciaria'

# Carregar a planilha Excel
workbook = openpyxl.load_workbook(nome_arquivo_excel)
planilha = workbook[nome_planilha_excel]

# Começando da linha 4, coluna 3 (C)
start_row = 4
column_index = 3

# Descobrir o número total de linhas na planilha
total_rows = planilha.max_row

# Loop para consultar todas as linhas da coluna 3, começando da linha 4
row = start_row  # Inicialize a variável row
primeiro_registro = True  # Variável para controlar o primeiro registro

solver = TwoCaptcha(Var_2captcha)

while row <= total_rows:
    Num_CDA = planilha.cell(row=row, column=column_index).value

    # Verifique se Num_CDA está vazio
    if Num_CDA is None:
        print("Num_CDA está vazio. Saindo do loop.")
        break  # Sai do loop se Num_CDA estiver vazio   
    
    # Digitar o CDA
    Campo_CDA = driver.find_element(By.XPATH, "//input[@id='consultaDebitoForm:decTxtTipoConsulta:cdaEtiqueta']")
    Campo_CDA.clear()  # Limpar o campo antes de inserir um novo valor
    Campo_CDA.send_keys(Num_CDA)
   
    # Se for o primeiro registro, espere 35 segundos
    if primeiro_registro:
        chave_captcha = driver.find_element(By.CLASS_NAME, 'g-recaptcha').get_attribute('data-sitekey')

        solver = recaptchaV2Proxyless()
        solver.set_verbose(1)
        solver.set_key(chave_api)
        solver.set_website_url(link)
        solver.set_website_key(chave_captcha)

        resposta = solver.solve_and_return_solution()

        if resposta != 0:
            print(resposta)
            # preencher o campo do token do captcha
            driver.execute_script(f"document.getElementById('g-recaptcha-response').innerHTML = '{resposta}'")
            driver.find_element(By.XPATH, "//*[@id='consultaDebitoForm:j_id78_body']/div[2]/input[2]").click()
        else:
            print(solver.err_string)

        primeiro_registro = False  # Defina como False após o primeiro registro

    # Verifique se a mensagem "Nenhum resultado com os critérios de consulta" está presente
    resultado_msg_element = driver.find_elements(By.XPATH, "//div[@class='rich-panel-body ']//p")
    resultado_msg = resultado_msg_element[0].text if resultado_msg_element else ""
    
    # Escreva a mensagem no Excel
    planilha.cell(row=row, column=column_index + 1, value=resultado_msg)
    if "Nenhum resultado com os critérios de consulta" in resultado_msg:
        CDA = str(Num_CDA)
        # Gerar PDF da Tela
        salvar_como_pdf(CDA)
        row += 1  # Vá para a próxima linha e continue o loop 
        continue
        
    # Consultar IPVA
    href_IPVA = driver.find_element(By.XPATH, "//*[@id='consultaDebitoForm:dataTable:0:lnkConsultaDebito']")
    href_IPVA.click()
    sleep(1)
    
    # Consultar Registro
    href_Registro = driver.find_element(By.XPATH, "//a[@href='#']")
    href_Registro.click()
    sleep(1)
     
    numero_processo_element = driver.find_element(By.XPATH, "//div[@id='consultaDebitoForm:j_id1058']")
    numero_processo = numero_processo_element.find_element(By.TAG_NAME, "span").text
    sleep(1)


    # Capturar a nova janela ou aba aberta
    janelas = driver.window_handles
    driver.switch_to.window(janelas[-1])
    sleep(2)
    
    # Extrair informações da nova janela
    situacao_protesto = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, "//*[@id='consultaDebitoForm:modalDadosCartorioContentTable']/tbody/tr[2]/td[1]"))
    ).text
    comarca_foro = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, "//*[@id='consultaDebitoForm:modalDadosCartorioContentTable']/tbody/tr[2]/td[2]"))
    ).text
    numero_protocolo = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, "//*[@id='consultaDebitoForm:modalDadosCartorioContentTable']/tbody/tr[2]/td[3]"))
    ).text
    data_abertura = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, "//*[@id='consultaDebitoForm:modalDadosCartorioContentTable']/tbody/tr[2]/td[4]"))
    ).text
    nome_cartorio = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, "//*[@id='consultaDebitoForm:tabelaDadosCartorio:0:j_id1300']"))
    ).text
    endereco_cartorio = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, "//*[@id='consultaDebitoForm:tabelaDadosCartorio:0:j_id1302']"))
    ).text
    bairro_cartorio = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, "//*[@id='consultaDebitoForm:j_id1072']"))
    ).text
    localidade_cartorio = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, "//*[@id='consultaDebitoForm:tabelaDadosCartorio:0:j_id1304']"))
    ).text
    CEP_cartorio = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, "//*[@id='consultaDebitoForm:tabelaDadosCartorio:0:j_id1308']"))
    ).text
    email = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, "//*[@id='consultaDebitoForm:tabelaDadosCartorio:0:j_id1310']"))
    ).text
    telefone_cartorio = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH,"//*[@id='consultaDebitoForm:tabelaDadosCartorio:0:j_id1312']"))
    ).text
    
    # Fechar a nova janela ou aba e voltar para a janela principal
    driver.close()
    driver.switch_to.window(janelas[0])
    sleep(2)
    
    # Encontre a próxima coluna disponível (vamos começar da coluna D)
    coluna_atual = column_index + 1

    # Escreva os valores nas colunas a partir da coluna D
    planilha.cell(row=row, column=coluna_atual, value=comarca_foro)
    coluna_atual += 1

    planilha.cell(row=row, column=coluna_atual, value=numero_processo)
    coluna_atual += 1

    planilha.cell(row=row, column=coluna_atual, value=numero_protocolo)
    coluna_atual += 1

    planilha.cell(row=row, column=coluna_atual, value=situacao_protesto)
    coluna_atual += 1

    planilha.cell(row=row, column=coluna_atual, value=data_abertura)
    coluna_atual += 1

    planilha.cell(row=row, column=coluna_atual, value=nome_cartorio)
    coluna_atual += 1

    planilha.cell(row=row, column=coluna_atual, value=endereco_cartorio)
    coluna_atual += 1

    planilha.cell(row=row, column=coluna_atual, value=bairro_cartorio)
    coluna_atual += 1

    planilha.cell(row=row, column=coluna_atual, value=localidade_cartorio)
    coluna_atual += 1

    planilha.cell(row=row, column=coluna_atual, value=CEP_cartorio)
    coluna_atual += 1

    planilha.cell(row=row, column=coluna_atual, value=email)
    coluna_atual += 1

    planilha.cell(row=row, column=coluna_atual, value=telefone_cartorio)
    coluna_atual += 1

    # Botão Voltar
    btn_Voltar1 = driver.find_element(By.XPATH, "//input[@name='consultaDebitoForm:j_id280']")
    btn_Voltar1.click()
    sleep(2)
    
    row += 1

# Salve o arquivo Excel atualizado
workbook.save(nome_arquivo_excel)

# Fechar o WebDriver
driver.quit()
