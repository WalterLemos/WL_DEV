from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.select import Select
import re
from twocaptcha import TwoCaptcha
import pyautogui
from time import sleep
import openpyxl
from Chave import *

def new_func():
    pyautogui.press('backspace', presses=20)

def salvar_como_pdf(renavam):

     pyautogui.hotkey('ctrl', 'p')
     sleep(2)
    # Mover e clicar na opção "Salvar como PDF"
    # Atualize as coordenadas de acordo com sua tela
    # clicar em guardar
     pyautogui.click(1519, 259)
     sleep(1)
     pyautogui.click(1477, 330)
     sleep(1)
     pyautogui.click(1493, 918)
     sleep(1)
     # apagar nome
     new_func()
     # Escreve a Renavam
     pyautogui.write(f'relatorio_{renavam}_{Num_Exercicio}_{Cota}.pdf')
     # clicar em salvar
     pyautogui.click(716, 670)
     sleep(1)
     # clicar em sim (caso tenha algo para substituir)
     pyautogui.click(1033, 527)    

# Inicializando o driver do Chrome
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
link = 'https://www.ipva.fazenda.sp.gov.br/IPVANET_CertidaoRecolhimento/'
driver.get(link)

pyautogui.click(1480, 39)
sleep(1)
pyautogui.click(1868, 184)
sleep(1)

# Nome do arquivo Excel e nome da planilha
nome_arquivo_excel = r'C:\Users\walter.oliveira\Documents\ProjetosPython\dev\Bichara_Dev\repository\Templete_Pagamentos_SP.xlsx'
nome_planilha_excel = 'Parcela 1'

# Carregar a planilha Excel
workbook = openpyxl.load_workbook(nome_arquivo_excel)
planilha = workbook[nome_planilha_excel]

# Começando da linha 4, coluna 3 (C)
start_row = 3
column_index = 2
column_index2 = 3
column_index3 = 4

# Descobrir o número total de linhas na planilha
total_rows = planilha.max_row

 # Configurar o solver do 2captcha
solver = TwoCaptcha(Var_2captcha)

  # Loop para consultar todas as linhas da coluna 3, começando da linha 4
row = start_row  # Inicialize a variável row

while row <= total_rows:
        Num_Exercicio = planilha.cell(row=row, column=column_index).value
        Num_Renavam = planilha.cell(row=row, column=column_index2).value
        Num_Placa = planilha.cell(row=row, column=column_index3).value

        # Certifique-se de que Num_Exercicio é uma string
        Num_Exercicio = str(Num_Exercicio).strip()
        Num_Renavam = str(Num_Renavam).strip()
        Num_Placa = str(Num_Placa).strip()

        # Esperar até que os campos estejam visíveis e habilitados
        campo_renavam = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.ID, "Renavam"))
        )
        campo_renavam.clear()  # Limpar o campo antes de inserir um novo valor
        campo_renavam.send_keys(Num_Renavam)

        campo_placa = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.ID, "Placa"))
        )
        campo_placa.clear()  # Limpar o campo antes de inserir um novo valor
        campo_placa.send_keys(Num_Placa)

        # Localizar o seletor pelo caminho especificado
        select_element = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable(
                (By.CSS_SELECTOR, "body > div.container.body-content > div:nth-child(7) > div > form > div:nth-child(3) > div > select"))
        )

        # Criar um objeto Select
        select = Select(select_element)

        # Selecionar o exercício na lista suspensa
        try:
            select.select_by_visible_text(Num_Exercicio)
        except Exception as e:
            print(f"Erro ao selecionar o exercício: {e}")
            sleep(2)
            row += 1
            continue

        chave_captcha = driver.find_element(By.CLASS_NAME, 'g-recaptcha').get_attribute('data-sitekey')

        # Resolver o captcha usando 2captcha
        try:
            result = solver.recaptcha(sitekey=chave_captcha, url=link)
            resposta = result['code']
        except Exception as e:
            print(f"Erro ao resolver o CAPTCHA: {e}")
            resposta = None

        if resposta:
            # Preencher o campo do token do captcha
            driver.execute_script(
                f"document.getElementById('g-recaptcha-response').innerHTML = '{resposta}'")

            # Esperar até que o botão esteja clicável
            try:
                btn_submit = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable(
                        (By.XPATH, "/html/body/div[1]/div[3]/div/form/div[7]/table/tbody/tr/td/input"))
                )
                btn_submit.click()
            except Exception as e:
                print(f"Erro ao clicar no botão de submissão: {e}")
                sleep(2)
                row += 1
                continue

            # Verificar se a mensagem "Dados não encontrados para o Renavam / placa / referência informados." está presente
            resultado_msg_element = driver.find_elements( By.XPATH, "/html/body/div[1]/div[5]/div/pre")
            resultado_msg = resultado_msg_element[0].text if resultado_msg_element else ""

            if "Dados não encontrados para o Renavam / placa / referência informados." in resultado_msg:
                # Escrever a mensagem no Excel
                planilha.cell(row=row, column=column_index3 +
                              1, value=resultado_msg)

                # Salvar a planilha após cada atualização
                workbook.save(nome_arquivo_excel)
                row += 1  # Vá para a próxima linha e continue o loop
                continue
            else:
                # Esperar que o elemento esteja presente
                wait = WebDriverWait(driver, 10)
                parcela1_element = wait.until(EC.presence_of_element_located((By.XPATH, '/html/body/div/div[8]/table/tbody/tr[4]/td[1]')))
                if parcela1_element.text == 'Parcela 1': 
                # Gerar Certidão
                   btn_Gerar = driver.find_element(By.XPATH, "/html/body/div/div[8]/table/tbody/tr[4]/td[2]/form")
                   btn_Gerar.click()

                # Extrair informações
                   num_controle_element = driver.find_element(By.XPATH, "//html/body/div/div[1]/div[4]/div[2]")
                   num_controle_texto = num_controle_element.text
                   num_controle = ''.join(re.findall(r'\d+', num_controle_texto))

                   ano_referencia_element = driver.find_element(By.XPATH, "/html/body/div/div[1]/div[4]/div[3]")
                   ano_referencia_texto = ano_referencia_element.text
                   ano_referencia = ''.join(re.findall(r'\d+', ano_referencia_texto))

                   data_arrecad_element = driver.find_element(By.XPATH, "/html/body/div/div[1]/div[4]/div[5]")
                   data_arrecad_texto = data_arrecad_element.text
                   data_arrecadacao = re.search(r'\d{2}/\d{2}/\d{4}', data_arrecad_texto).group()

                   Cota_element = driver.find_element(By.XPATH, "/html/body/div/div[1]/div[4]/div[4]")
                   Cota_texto = Cota_element.text.strip()
                   Cota = Cota_texto.split("Cota:")[1].strip()

                   banco_agencia_element = driver.find_element(By.XPATH, "/html/body/div/div[1]/div[4]/div[6]")
                   banco_agencia_texto = banco_agencia_element.text.strip()
                   banco_agencia = banco_agencia_texto.split("Banco/Agência:")[1].strip()

                   valor_recolhido_element = driver.find_element(By.XPATH, "/html/body/div/div[1]/div[5]/div[4]")
                   valor_recolhido_texto = valor_recolhido_element.text.strip()
                   valor_recolhido = valor_recolhido_texto.split("R$")[1].strip()

                   valor_restituido_element = driver.find_element(By.XPATH, "/html/body/div/div[1]/div[5]/div[6]")
                   valor_restituido_texto = valor_restituido_element.text.strip()
                   valor_restituido = valor_restituido_texto.split("R$")[1].strip()

                   valor_total_element = driver.find_element(By.XPATH, "/html/body/div/div[1]/div[5]/div[8]")
                   valor_total_texto = valor_total_element.text.strip()
                   valor_total = valor_total_texto.split("R$")[1].strip()

                  # Encontre a próxima coluna disponível (vamos começar da coluna E)
                   coluna_atual = column_index3 + 1

                 # Escreva os valores nas colunas a partir da coluna E
                   planilha.cell(row=row, column=coluna_atual, value=num_controle)
                   coluna_atual += 1

                   planilha.cell(row=row, column=coluna_atual, value=ano_referencia)
                   coluna_atual += 1

                   planilha.cell(row=row, column=coluna_atual,value=data_arrecadacao)
                   coluna_atual += 1

                   planilha.cell(row=row, column=coluna_atual, value=Cota)
                   coluna_atual += 1

                   planilha.cell(row=row, column=coluna_atual,value=banco_agencia)
                   coluna_atual += 1

                   planilha.cell(row=row, column=coluna_atual, value=valor_recolhido)
                   coluna_atual += 1

                   planilha.cell(row=row, column=coluna_atual, value=valor_restituido)
                   coluna_atual += 1

                   planilha.cell(row=row, column=coluna_atual, value=valor_total)
                   coluna_atual += 1

                   # Salvar a planilha após cada atualização
                   workbook.save(nome_arquivo_excel)
                   # Manipular a tela de impressão para salvar em PDF
                   renavam = str(Num_Renavam)
                   salvar_como_pdf(renavam)

                   btn_retornar = driver.find_element(By.XPATH, '/html/body/div/div[3]/div[2]/button')
                   btn_retornar.click()
                   sleep(2)
                   continue    
                # Verificar se o texto do elemento é 'Parcela Unica'
                parcela_unica_element = driver.find_element(By.XPATH,'/html/body/div/div[8]/table/tbody/tr[5]/td[1]')
                parcela_unica = parcela_unica_element.text  
        
                if parcela_unica == 'Parcela Única':
                    # Nome do arquivo Excel e nome da planilha
                      nome_arquivo_excel = r'C:\Users\walter.oliveira\Documents\ProjetosPython\dev\Bichara_Dev\repository\Templete_Pagamentos_SP.xlsx'
                      nome_planilha_excel = 'Parcela única'

                   # Carregar a planilha Excel
                      workbook = openpyxl.load_workbook(nome_arquivo_excel)
                      planilha = workbook[nome_planilha_excel]

                      btn_gerar_parcela_unica = driver.find_element(By.XPATH, '/html/body/div/div[8]/table/tbody/tr[4]/td[2]/form' )            
                      btn_gerar_parcela_unica.click()

                    # Extrair informações
                      num_controle_parcela_element = driver.find_element(By.XPATH, "/html/body/div/div[1]/div[4]/div[2]")
                      num_controle_parcela_texto =  num_controle_parcela_element.text
                      num_controle_parcela = ''.join(re.findall(r'\d+', num_controle_parcela_texto))

                      ano_referencia_parcela_element = driver.find_element(By.XPATH, "/html/body/div/div[1]/div[4]/div[3]")
                      ano_referencia_parcela_texto = ano_referencia_parcela_element.text
                      ano_referencia_parcela = ''.join(re.findall(r'\d+', ano_referencia_parcela_texto))

                      data_arrecad_parcela_element = driver.find_element(By.XPATH, "/html/body/div/div[1]/div[4]/div[5]")
                      data_arrecad_parcela_texto = data_arrecad_parcela_element.text
                      data_arrecadacao_parcela = re.search(r'\d{2}/\d{2}/\d{4}', data_arrecad_parcela_texto).group()

                      Cota_parcela_element = driver.find_element(By.XPATH, "/html/body/div/div[1]/div[4]/div[4]")
                      Cota_parcela_texto = Cota_parcela_element.text.strip()
                      Cota_parcela = Cota_parcela_texto.split("Cota:")[1].strip()

                      banco_agencia_parcela_element = driver.find_element(By.XPATH, "/html/body/div/div[1]/div[4]/div[6]")
                      banco_agencia_parcela_texto = banco_agencia_parcela_element.text.strip()
                      banco_agencia_parcela = banco_agencia_parcela_texto.split("Banco/Agência:")[1].strip()

                      valor_recolhido_parcela_element = driver.find_element(By.XPATH, "/html/body/div/div[1]/div[5]/div[4]")
                      valor_recolhido_parcela_texto = valor_recolhido_parcela_element.text.strip()
                      valor_recolhido_parcela = valor_recolhido_parcela_texto.split("R$")[1].strip()

                      valor_restituido_parcela_element = driver.find_element(By.XPATH, "/html/body/div/div[1]/div[5]/div[6]")
                      valor_restituido_parcela_texto = valor_restituido_parcela_element.text.strip()
                      valor_restituido_parcela = valor_restituido_parcela_texto.split("R$")[1].strip()

                      valor_total_parcela_element = driver.find_element(By.XPATH, "/html/body/div/div[1]/div[5]/div[8]")
                      valor_total_parcela_texto = valor_total_parcela_element.text.strip()
                      valor_total_parcela = valor_total_parcela_texto.split("R$")[1].strip()

                      # Encontre a próxima coluna disponível (vamos começar da coluna E)
                      coluna_atual = column_index3 + 1

                      # Escreva os valores nas colunas a partir da coluna E
                      planilha.cell(row=row, column=coluna_atual, value=num_controle_parcela)
                      coluna_atual += 1

                      planilha.cell(row=row, column=coluna_atual, value=ano_referencia_parcela)
                      coluna_atual += 1

                      planilha.cell(row=row, column=coluna_atual, value=data_arrecadacao_parcela)
                      coluna_atual += 1

                      planilha.cell(row=row, column=coluna_atual, value=Cota_parcela)
                      coluna_atual += 1

                      planilha.cell(row=row, column=coluna_atual, value=banco_agencia_parcela)
                      coluna_atual += 1

                      planilha.cell(row=row, column=coluna_atual, value=valor_recolhido_parcela)
                      coluna_atual += 1

                      planilha.cell(row=row, column=coluna_atual, value=valor_restituido_parcela)
                      coluna_atual += 1

                      planilha.cell(row=row, column=coluna_atual, value=valor_total_parcela)
                      coluna_atual += 1

                      workbook.save(nome_arquivo_excel)

                      salvar_como_pdf(renavam)

                      btn_Voltar_parcela = driver.find_element(By.XPATH, '/html/body/div/div[3]/div[2]/button')
                      btn_Voltar_parcela.click()
                # Botão Retornar Pagina Inicial
                      btn_retornar_inicial = driver.find_element(By.XPATH, '/html/body/div/div[10]/div/button')
                      btn_retornar_inicial.click()
                      sleep(2)
                      row += 1
        else:
           print("Falha ao resolver o CAPTCHA.")
        row += 1
workbook.save(nome_arquivo_excel)
    # Fechar o navegador
driver.quit()
