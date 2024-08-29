from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.select import Select
import re
from twocaptcha import TwoCaptcha
import pyautogui
from time import sleep
import openpyxl
from Chave import *

# Inicializando o driver do Chrome
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
link = 'https://www.ipva.fazenda.sp.gov.br/IPVANET_CertidaoRecolhimento/'
driver.get(link)

# Ações usando pyautogui para ajustes na tela
pyautogui.click(1480, 39)
sleep(1)
pyautogui.click(1868, 184)
sleep(1)

# Carregar a planilha do Excel
nome_arquivo_excel = r'C:\Users\walter.oliveira\Documents\ProjetosPython\dev\Bichara_Dev\repository\Templete_Pagamentos_SP.xlsx'
workbook = openpyxl.load_workbook(nome_arquivo_excel)
planilha = workbook.active
total_rows = planilha.max_row

start_row = 3
column_index = 2
column_index2 = 3
column_index3 = 4

def new_func():
    pyautogui.press('backspace', presses=20)

def salvar_como_pdf(renavam, num_exercicio, cota):
    pyautogui.hotkey('ctrl', 'p')
    sleep(2)
    pyautogui.click(1519, 259)
    sleep(1)
    pyautogui.click(1477, 330)
    sleep(1)
    pyautogui.click(1493, 918)
    sleep(1)
    new_func()
    pyautogui.write(f'relatorio_{renavam}_{num_exercicio}_{cota}.pdf')
    pyautogui.click(716, 670)
    sleep(1)
    pyautogui.click(1033, 527)

def processar_parcela(parcela_texto, nome_planilha_excel, renavam, exercicio):
    global row
    planilha = workbook[nome_planilha_excel]
    
    # Verificação de elementos antes da interação
    try:
        num_controle_element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//html/body/div/div[1]/div[4]/div[2]"))
        )
        num_controle_texto = num_controle_element.text
        num_controle = ''.join(re.findall(r'\d+', num_controle_texto))

        ano_referencia_element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "/html/body/div/div[1]/div[4]/div[3]"))
        )
        ano_referencia_texto = ano_referencia_element.text
        ano_referencia = ''.join(re.findall(r'\d+', ano_referencia_texto))

        data_arrecad_element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "/html/body/div/div[1]/div[4]/div[5]"))
        )
        data_arrecad_texto = data_arrecad_element.text
        data_arrecadacao = re.search(r'\d{2}/\d{2}/\d{4}', data_arrecad_texto).group()

        Cota_element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "/html/body/div/div[1]/div[4]/div[4]"))
        )
        Cota_texto = Cota_element.text.strip()
        Cota = Cota_texto.split("Cota:")[1].strip()

        banco_agencia_element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "/html/body/div/div[1]/div[4]/div[6]"))
        )
        banco_agencia_texto = banco_agencia_element.text.strip()
        banco_agencia = banco_agencia_texto.split("Banco/Agência:")[1].strip()

        valor_recolhido_element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "/html/body/div/div[1]/div[5]/div[4]"))
        )
        valor_recolhido_texto = valor_recolhido_element.text.strip()
        valor_recolhido = valor_recolhido_texto.split("R$")[1].strip()

        valor_restituido_element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "/html/body/div/div[1]/div[5]/div[6]"))
        )
        valor_restituido_texto = valor_restituido_element.text.strip()
        valor_restituido = valor_restituido_texto.split("R$")[1].strip()

        valor_total_element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "/html/body/div/div[1]/div[5]/div[8]"))
        )
        valor_total_texto = valor_total_element.text.strip()
        valor_total = valor_total_texto.split("R$")[1].strip()
    except Exception as e:
        print(f"Erro ao processar elementos: {e}")
        return

    # Escreve os dados na planilha
    coluna_atual = column_index3 + 1

    planilha.cell(row=row, column=coluna_atual, value=num_controle)
    coluna_atual += 1
    planilha.cell(row=row, column=coluna_atual, value=ano_referencia)
    coluna_atual += 1
    planilha.cell(row=row, column=coluna_atual, value=data_arrecadacao)
    coluna_atual += 1
    planilha.cell(row=row, column=coluna_atual, value=Cota)
    coluna_atual += 1
    planilha.cell(row=row, column=coluna_atual, value=banco_agencia)
    coluna_atual += 1
    planilha.cell(row=row, column=coluna_atual, value=valor_recolhido)
    coluna_atual += 1
    planilha.cell(row=row, column=coluna_atual, value=valor_restituido)
    coluna_atual += 1
    planilha.cell(row=row, column=coluna_atual, value=valor_total)
    coluna_atual += 1
    workbook.save(nome_arquivo_excel)
    salvar_como_pdf(renavam, exercicio, Cota)
    
# Configurar o solver do 2captcha
solver = TwoCaptcha(Var_2captcha)

# Loop para consultar todas as linhas da coluna 3, começando da linha 4
row = start_row

while row <= total_rows:
    try:
        Num_Exercicio = planilha.cell(row=row, column=column_index).value
        Num_Renavam = planilha.cell(row=row, column=column_index2).value
        Num_Placa = planilha.cell(row=row, column=column_index3).value

        Num_Exercicio = str(Num_Exercicio).strip()
        Num_Renavam = str(Num_Renavam).strip()
        Num_Placa = str(Num_Placa).strip()

        campo_renavam = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.ID, "Renavam"))
        )
        campo_renavam.clear()
        campo_renavam.send_keys(Num_Renavam)

        campo_placa = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.ID, "Placa"))
        )
        campo_placa.clear()
        campo_placa.send_keys(Num_Placa)

        select_element = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, "body > div.container.body-content > div:nth-child(7) > div > form > div:nth-child(3) > div > select"))
        )
        select = Select(select_element)

        try:
            select.select_by_visible_text(Num_Exercicio)
        except Exception as e:
            print(f"Erro ao selecionar o exercício: {e}")
            sleep(2)
            row += 1
            continue

        chave_captcha = driver.find_element(By.CLASS_NAME, 'g-recaptcha').get_attribute('data-sitekey')

        try:
            result = solver.recaptcha(sitekey=chave_captcha, url=link)
            resposta = result['code']
        except Exception as e:
            print(f"Erro ao resolver o CAPTCHA: {e}")
            resposta = None

        if resposta:
            driver.execute_script(
                f"document.getElementById('g-recaptcha-response').innerHTML = '{resposta}'")

            btn_submit = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, "/html/body/div[1]/div[3]/div/form/div[7]/table/tbody/tr/td/input"))
            )
            btn_submit.click()

            resultado_msg_element = WebDriverWait(driver, 10).until(
            EC.presence_of_all_elements_located((By.XPATH, "/html/body/div[1]/div[5]/div/pre"))
            )
        resultado_msg = resultado_msg_element[0].text if resultado_msg_element else ""

        if "Dados não encontrados para o Renavam / placa / referência informados." in resultado_msg:
            planilha.cell(row=row, column=column_index3 + 1, value=resultado_msg)
            workbook.save(nome_arquivo_excel)
            row += 1
            continue
        else:
            try:
                parcela_element_1 = driver.find_element(By.XPATH, "//td[text()='Parcela 1']")
                parcela_element_unica = driver.find_element(By.XPATH, "//td[text()='Parcela Única']")
                parcela_unica = parcela_element_unica.text if parcela_element_unica else ""

                if parcela_element_1 and parcela_unica:
                    btn_Gerar = WebDriverWait(driver, 10).until(
                        EC.element_to_be_clickable((By.XPATH, "/html/body/div/div[8]/table/tbody/tr[4]/td[2]/form"))
                    )
                    btn_Gerar.click()
                    sleep(2)
                    processar_parcela(parcela_element_1.text, 'Parcela 1', Num_Renavam, Num_Exercicio)

                    btn_retornar = WebDriverWait(driver, 10).until(
                        EC.element_to_be_clickable((By.XPATH, '/html/body/div/div[3]/div[2]/button'))
                    )
                    btn_retornar.click()
                    sleep(2)

                    btn_gerar_parcela_unica = WebDriverWait(driver, 10).until(
                        EC.element_to_be_clickable((By.XPATH, '/html/body/div/div[8]/table/tbody/tr[4]/td[2]/form'))
                    )
                    btn_gerar_parcela_unica.click()
                    sleep(1)
                    processar_parcela(parcela_element_unica.text, 'Parcela única', Num_Renavam, Num_Exercicio)

                    btn_retornar = WebDriverWait(driver, 10).until(
                        EC.element_to_be_clickable((By.XPATH, '/html/body/div/div[3]/div[2]/button'))
                    )
                    btn_retornar.click()
                    sleep(2)
                    row += 1
                    continue

                if parcela_element_1:
                    btn_Gerar = WebDriverWait(driver, 10).until(
                        EC.element_to_be_clickable((By.XPATH, "/html/body/div/div[8]/table/tbody/tr[4]/td[2]/form"))
                    )
                    btn_Gerar.click()
                    sleep(2)
                    processar_parcela(parcela_element_1.text, 'Parcela 1', Num_Renavam, Num_Exercicio)

                    btn_retornar = WebDriverWait(driver, 10).until(
                        EC.element_to_be_clickable((By.XPATH, '/html/body/div/div[3]/div[2]/button'))
                    )
                    btn_retornar.click()
                    sleep(2)
                    row += 1
                    continue

                if parcela_unica:
                    btn_Gerar = WebDriverWait(driver, 10).until(
                        EC.element_to_be_clickable((By.XPATH, "/html/body/div/div[8]/table/tbody/tr[4]/td[2]/form"))
                    )
                    btn_Gerar.click()
                    sleep(2)
                    processar_parcela(parcela_element_unica.text, 'Parcela única', Num_Renavam, Num_Exercicio)

                    btn_retornar = WebDriverWait(driver, 10).until(
                        EC.element_to_be_clickable((By.XPATH, '/html/body/div/div[3]/div[2]/button'))
                    )
                    btn_retornar.click()
                    sleep(2)
                    row += 1
                    continue

            except Exception as e:
                print(f"Erro ao processar parcelas: {e}")
                row += 1
                continue

    except Exception as e:
        print(f"Erro geral: {e}")
        row += 1

# Botão Retornar Página Inicial
try:
    btn_retornar_inicial = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.XPATH, '/html/body/div/div[10]/div/button'))
    )
    btn_retornar_inicial.click()
    sleep(2)
except Exception as e:
    print(f"Erro ao retornar à página inicial: {e}")

# Encerrar o driver
driver.quit()

