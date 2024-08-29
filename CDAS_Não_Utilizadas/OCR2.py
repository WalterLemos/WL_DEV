import openpyxl
import os
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
from time import sleep

# Configurações iniciais
numero_oab = 88922
chrome_driver_path = r'C:\ProjetorPython\dev\chromedriver-win64\chromedriver.exe'
os.environ["webdriver.chrome.driver"] = chrome_driver_path
driver = webdriver.Chrome()

# Função para criar uma nova sheet no workbook
def create_sheet(workbook, sheet_name):
    try:
        sheet = workbook[sheet_name]
    except KeyError:
        sheet = workbook.create_sheet(sheet_name)
    return sheet

# Função para adicionar headers na sheet
def add_headers(sheet):
    sheet['A1'].value = "Número Processo"
    sheet['B1'].value = "Data Distribuição"
    sheet['C1'].value = "Movimentações"

# Função para adicionar dados na sheet
def add_data(sheet, numero_processo, data_distribuicao, lista_movimentacoes):
    sheet['A2'].value = numero_processo
    sheet['B2'].value = data_distribuicao
    for index, row in enumerate(sheet.iter_rows(min_row=2, max_row=len(lista_movimentacoes)+1, min_col=3, max_col=3)):
        for cell in row:
            cell.value = lista_movimentacoes[index]

# Main function
def main():
    # Acessar pagina da consulta pública
    driver.get('https://pje-consulta-publica.tjmg.jus.br/')
    sleep(10)
    
    # Digitar número OAB
    campo_oab = driver.find_element(By.XPATH, "//input[@id='fPP:Decoration:numeroOAB']")
    campo_oab.send_keys(numero_oab)
    
    # Selecionar estado
    dropdown_estados = driver.find_element(By.XPATH, "//select[@id='fPP:Decoration:estadoComboOAB']")
    opcoes_estados = Select(dropdown_estados)
    opcoes_estados.select_by_visible_text('RJ')
    
    # Clicar em pesquisar
    botao_pesquisar = driver.find_element(By.XPATH, "//input[@id='fPP:searchProcessos']")
    botao_pesquisar.click()
    sleep(10)
    
    # Entrar em cada um dos processos
    processos = driver.find_elements(By.XPATH, "//b[@class='btn-block']")
    for proceso in processos:
        proceso.click()
        sleep(10)
        janelas = driver.window_handles
        driver.switch_to.window(janelas[-1])
        driver.set_window_size(1920, 1080)
        
        # Extrair o n° do processo e data da distribuição
        numero_processo = driver.find_elements(By.XPATH, "//div[@class='col-sm-12 ']")
        numero_processo = numero_processo[0].text
        
        data_distribuicao = driver.find_elements(By.XPATH, "//div[@class='value col-sm-12 ']")
        data_distribuicao = data_distribuicao[1].text
        
        # Extrair e guardar todas as últimas movimentações
        movimentacoes = driver.find_elements(By.XPATH, "//div[@id='j_id132:processoEventoPanel_body']//tr[contains(@class,'rich-table-row')]//td//div//div//span")
        lista_movimentacoes = [movimentacao.text for movimentacao in movimentacoes]
        
        # Guardar tudo no excel, separados por processo
        workbook = openpyxl.load_workbook(r'C:/Users/walter.oliveira/Documents/ProjetosPython/dev/Bichara_Dev/py/dados.xlsx')
        sheet_name = numero_processo
        try:
            sheet = workbook[sheet_name]
        except KeyError:
            sheet = create_sheet(workbook, sheet_name)
            add_headers(sheet)
        
        add_data(sheet, numero_processo, data_distribuicao, lista_movimentacoes)
        workbook.save('dados.xlsx')
        
        driver.close()
        sleep(5)
        driver.switch_to.window(driver.window_handles[0])

if __name__ == '__main__':
    main()