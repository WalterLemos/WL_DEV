from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.select import Select
from selenium.common.exceptions import TimeoutException
import pyautogui
from time import sleep
import openpyxl
from Chave import *

def new_func():
    pyautogui.press('backspace', presses=20)

def Consultar_CNPJ():

     # Digitar a Raiz do CNPJ
    CNPJ = str(Num_CNPJ)
    pyautogui.click(484,451)
    pyautogui.write(CNPJ)
    sleep(3)
   #Selecinar o Impeditivo de CND
    pyautogui.click(567,495)
    sleep(3)
    pyautogui.click(524,588)
    sleep(3)
    #Clicar em Consultar
    pyautogui.click(80,300)  
    sleep(5)
    #Selecionar Registro
    pyautogui.click(428,928)
    sleep(2)
    #Clicar em Pagar
    pyautogui.click(389,296)
    sleep(2)
    #Selecionar Boleto
    pyautogui.click(929,666)
    sleep(2)
    #Clicar em Pagar
    pyautogui.click(890,882)
    sleep(2)

def salvar_como_pdf(guia):
    pyautogui.click(1817,316)
    sleep(2)
    pyautogui.click(1519, 259)
    sleep(1)
    pyautogui.click(1477, 330)
    sleep(1)
    pyautogui.click(1493, 918)
    sleep(1)
    new_func()
    pyautogui.write(f'Subguia_Guia_{guia}.pdf')
    pyautogui.click(716, 670)
    sleep(1)
    pyautogui.click(1033, 527)

# Inicializando o driver do Chrome
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
link = 'https://www.tjsc.jus.br/consulta-e-pagamento-de-custas-e-outros-debitos'
driver.get(link)

# Esperar e clicar nos elementos usando pyautogui
pyautogui.click(1474, 44)
sleep(3)
pyautogui.click(1870, 180)
sleep(3)

# Esperar até que o botão "Consultar" esteja presente
btn_Consultar = WebDriverWait(driver, 10).until(
    EC.element_to_be_clickable((By.XPATH, "//*[@id='portlet_com_liferay_journal_content_web_portlet_JournalContentPortlet_INSTANCE_97S3MWFSrQvk']/div/div[2]/div/div/div/ul/li/a"))
)
btn_Consultar.click()
sleep(3)

# Clicar em outro elemento usando pyautogui
pyautogui.click(124, 365)
sleep(2)

# Nome do arquivo Excel e nome da planilha
nome_arquivo_excel = r'C:\Users\walter.oliveira\Documents\ProjetosPython\dev\Bichara_Dev\repository\Templete_Taxa_Judiciaria_SC.rar.xlsx'
nome_planilha_excel = 'Débitos Taxa Judiciaria'

# Carregar a planilha Excel
workbook = openpyxl.load_workbook(nome_arquivo_excel)
planilha = workbook[nome_planilha_excel]

# Começando da linha 4, coluna 2 (B)
start_row = 4
column_index = 2

# Descobrir o número total de linhas na planilha
total_rows = planilha.max_row

# Loop para consultar todas as linhas da coluna 3, começando da linha 4
row = start_row  # Inicialize a variável row
primeiro_registro = True  # Variável para controlar o primeiro registro

#solver = TwoCaptcha(Var_2captcha)

while row <= total_rows:
    Num_CNPJ = planilha.cell(row=row, column=column_index).value

    # Verifique se Num_CDA está vazio
    if Num_CNPJ is None:
        print("Num_CDA está vazio. Saindo do loop.")
        break  # Sai do loop se Num_CDA estiver vazio

    Consultar_CNPJ() 
    sleep(10)
    try:
     # Wait for the element to be present
      guia_element = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, "//*[@id='form:consultaDividasCustasDataTable:dataTable:0:n']"))
        )
    # Perform actions on the element
      print("Element found:", guia_element)
    except Exception as e:
      print("Error:", e)
    subguia_guia = guia_element.find_element(By.TAG_NAME, 'span').text
    salvar_como_pdf(subguia_guia)
    sleep(3)  # Adicionar um pequeno atraso após clicar no botão

    row += 1  # Incrementar a linha para o próximo registro

# Fechar o driver após finalizar o processo
driver.quit()
