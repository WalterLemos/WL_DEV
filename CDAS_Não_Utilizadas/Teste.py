from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.select import Select
import re
from twocaptcha import TwoCaptcha
from anticaptchaofficial.recaptchav2proxyless import *
import pyautogui
from time import sleep
import openpyxl
from Chave import *
from bs4 import BeautifulSoup

def new_func():
    pyautogui.press('backspace', presses=20)

def salvar_como_pdf(CDA, processo):
    pyautogui.click(1777,93)
    sleep(3)
    pyautogui.click(1442,227)
    sleep(3)
    pyautogui.click(1817,170)
    sleep(3)
    pyautogui.click(1549,260)
    sleep(3)
    pyautogui.click(1513,332)
    sleep(3)
    pyautogui.click(1482,915)
    sleep(3)
    new_func()
    sleep(3)
    pyautogui.write(f'relatorio_{CDA}_{processo}.pdf')
    sleep(3)
    pyautogui.click(702,668)
    sleep(3)
    pyautogui.click(1031,530)

# Inicializando o driver do Chrome
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
link = 'https://www.dividaativa.pge.sp.gov.br/sc/pages/consultas/consultarDebito.jsf'
driver.get(link)

pyautogui.click(1474,44)
sleep(2)
pyautogui.click(1870,180)
sleep(2)



# Nome do arquivo Excel e nome da planilha
nome_arquivo_excel = r'C:\Users\walter.oliveira\Documents\ProjetosPython\dev\Bichara_Dev\repository\Templete_Taxa_Judiciaria.xlsx'
nome_planilha_excel = 'Débitos Taxa Judiciaria'

# Carregar a planilha Excel
workbook = openpyxl.load_workbook(nome_arquivo_excel)
planilha = workbook[nome_planilha_excel]

# Começando da linha 4, coluna 3 (C)
start_row = 4
column_index = 3

# Descobrir o número total de linhas na planilha
total_rows = planilha.max_row

# Loop para consultar todas as linhas da coluna 3, começando da linha 4
row = start_row  # Inicialize a variável row
primeiro_registro = True  # Variável para controlar o primeiro registro

#solver = TwoCaptcha(Var_2captcha)

while row <= total_rows:
    Num_CDA = planilha.cell(row=row, column=column_index).value

    # Verifique se Num_CDA está vazio
    if Num_CDA is None:
        print("Num_CDA está vazio. Saindo do loop.")
        break  # Sai do loop se Num_CDA estiver vazio   
    
    # Digitar o CDA
    Campo_CDA = driver.find_element(By.XPATH, "//input[@id='consultaDebitoForm:decTxtTipoConsulta:cdaEtiqueta']")
    Campo_CDA.clear()  # Limpar o campo antes de inserir um novo valor
    Campo_CDA.send_keys(Num_CDA)
   
    # Se for o primeiro registro, espere 35 segundos
    if primeiro_registro:
       # chave_captcha = driver.find_element(By.CLASS_NAME, 'g-recaptcha').get_attribute('data-sitekey')

        #solver = recaptchaV2Proxyless()
        #solver.set_verbose(1)
        #solver.set_key(chave_api)
        #solver.set_website_url(link)
        #solver.set_website_key(chave_captcha)

        #resposta = solver.solve_and_return_solution()

       # if resposta != 0:
           # print(resposta)
            # preencher o campo do token do captcha
            #driver.execute_script(f"document.getElementById('g-recaptcha-response').innerHTML = '{resposta}'")
           # driver.find_element(By.XPATH, "//*[@id='consultaDebitoForm:j_id78_body']/div[2]/input[2]").click()
       # else:
            #print(solver.err_string)
        sleep(35)
        primeiro_registro = False  # Defina como False após o primeiro registro

    # Verifique se a mensagem "Nenhum resultado com os critérios de consulta" está presente
    resultado_msg_element = driver.find_elements(By.XPATH, "//div[@class='rich-panel-body ']//p")
    resultado_msg = resultado_msg_element[0].text if resultado_msg_element else ""
    
    # Escreva a mensagem no Excel
    planilha.cell(row=row, column=column_index + 1, value=resultado_msg)
    if "Nenhum resultado com os critérios de consulta" in resultado_msg:
        CDA = str(Num_CDA)
        # Gerar PDF da Tela
        salvar_como_pdf(CDA)
        row += 1  # Vá para a próxima linha e continue o loop 
        continue
        
    # Consultar IPVA
    href_IPVA = driver.find_element(By.XPATH, "//*[@id='consultaDebitoForm:dataTable:0:lnkConsultaDebito']")
    href_IPVA.click()
    sleep(1)

    btn_liquidar = driver.find_element(By.XPATH, "//*[@id='consultaDebitoForm:dataTable2:0:lnkLiquidarDebito']" )
    btn_liquidar.click()
    sleep(2)

    numero_processo_element = driver.find_element(By.XPATH, "//*[@id='consultaDebitoForm:j_id369']")
    numero_processo = numero_processo_element.find_element(By.TAG_NAME, "span").text

    # Verificar se o elemento está visível e clicável
    btn_Gerar_Gare = WebDriverWait(driver, 10).until(
        EC.visibility_of_element_located((By.XPATH, "//*[@id='consultaDebitoForm:btnDownload']"))
    )
    btn_Gerar_Gare = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.XPATH, "//*[@id='consultaDebitoForm:btnDownload']"))
    )
    btn_Gerar_Gare.click()
    sleep(2)

    pyautogui.click(1348,450)
    sleep(2)
    # Esperar até que o painel modal desapareça
    WebDriverWait(driver, 10).until(
        EC.invisibility_of_element((By.ID, "modalPanelMsgGerarGareDiv"))
    )

    btn_download = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.XPATH, "//*[@id='consultaDebitoForm:btnDownloadGare']"))
    )
    btn_download.click()
    sleep(2)

    salvar_como_pdf(Num_CDA, numero_processo) 
    sleep(2)

    # Botão Voltar
    btn_Voltar = driver.find_element(By.XPATH, "//*[@id='consultaDebitoForm:j_id854_body']/div[1]/input[1]")
    btn_Voltar.click()
    sleep(2)
    
    # Botão Voltar
    btn_Voltar1 = driver.find_element(By.XPATH, "//*[@id='consultaDebitoForm:btnVoltar']")
    btn_Voltar1.click()
    sleep(2)

    # Botão Voltar
    btn_Voltar2 = driver.find_element(By.XPATH, "//*[@id='consultaDebitoForm:consultaDebito']/div[2]/input")
    btn_Voltar2.click()
    sleep(2)
     
    row += 1

# Salve o arquivo Excel atualizado
workbook.save(nome_arquivo_excel)

# Fechar o WebDriver
driver.quit()