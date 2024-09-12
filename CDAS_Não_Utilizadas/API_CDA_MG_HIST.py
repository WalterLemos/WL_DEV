from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException
import pyautogui
from time import sleep
import openpyxl

def new_func():
    pyautogui.press('backspace',presses=20)

# entrar no site da -https://www.dividaativa.pge.sp.gov.br/sc/pages/consultas/consultarDebito.jsf
chrome_driver_path = r'C:\ProjetorPython\chromedriver-win64\chromedriver-win64\chromedriver.exe'
driver = webdriver.Chrome(executable_path=chrome_driver_path)
driver.get('https://www2.fazenda.mg.gov.br/sol/ctrl/SOL/DIVATIV/SERVICO_001?ACAO=IMPRIMIR')

pyautogui.click(982,25)

# Nome do arquivo Excel e nome da planilha
nome_arquivo_excel = 'C:\\ProjetorPython\\Análise das CDAs-BV.xlsx'
nome_planilha_excel = '5272314-91.2022.8.13.0024'

# Carregar a planilha Excel
workbook = openpyxl.load_workbook(nome_arquivo_excel)
planilha = workbook[nome_planilha_excel]

# Começando da linha 2, coluna 3 (C)
start_row = 2
column_index = 3

# Começando da linha 2, coluna 2 (B)
start_row2 = 2
column_index2 = 2

# Descobrir o número total de linhas na planilha
total_rows = planilha.max_row

# Loop para consultar todas as linhas da coluna 3, começando da linha 4
row = start_row  # Inicialize a variável row
row2 =start_row2 
primeiro_registro = True  # Variável para controlar o primeiro registro

while row <= total_rows:
    Num_CDA = planilha.cell(row=row, column=column_index).value
    Num_ID = planilha.cell(row=row2, column=column_index2).value

# Digitar o CDA
    Campo_CDA = driver.find_element(By.XPATH, "//input[@name='txtNumeroPTA']")
    Campo_CDA.clear()  # Limpar o campo antes de inserir um novo valor
    Campo_CDA.send_keys(Num_CDA)
    sleep(40)
# Botão Emitir
    btn_emitir = driver.find_element(By.XPATH, "//img[@name='Emitir']")
    btn_emitir.click()
    sleep(5)
    pyautogui.click(795,96)
    sleep(2)
    pyautogui.click(146,376)
    sleep(2)
    pyautogui.write(Num_CDA)
    sleep(2)
    pyautogui.click(519,445)

   

# Fechar o WebDriver
driver.quit()   

  

