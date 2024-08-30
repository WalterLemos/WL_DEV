from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException
import pyautogui
from time import sleep
import openpyxl

def new_func():
    pyautogui.press('backspace',presses=20)

# entrar no site da -https://www.dividaativa.pge.sp.gov.br/sc/pages/consultas/consultarDebito.jsf
chrome_driver_path = r'C:\ProjetorPython\chromedriver-win64\chromedriver-win64\chromedriver.exe'
driver = webdriver.Chrome(executable_path=chrome_driver_path)
driver.get('http://receitaonline.fazenda.mg.gov.br/rol/dae/')

pyautogui.click(982,25)

# Nome do arquivo Excel e nome da planilha
nome_arquivo_excel = 'C:\\ProjetorPython\\Análise das CDAs-BV.xlsx'
nome_planilha_excel = '5272314-91.2022.8.13.0024'

# Carregar a planilha Excel
workbook = openpyxl.load_workbook(nome_arquivo_excel)
planilha = workbook[nome_planilha_excel]

# Começando da linha 2, coluna 3 (C)
start_row = 2
column_index = 3

# Começando da linha 2, coluna 2 (B)
start_row2 = 2
column_index2 = 2

# Descobrir o número total de linhas na planilha
total_rows = planilha.max_row

# Loop para consultar todas as linhas da coluna 3, começando da linha 4
row = start_row  # Inicialize a variável row
row2 =start_row2 
primeiro_registro = True  # Variável para controlar o primeiro registro

while row <= total_rows:
    Num_CDA = planilha.cell(row=row, column=column_index).value
    Num_ID = planilha.cell(row=row2, column=column_index2).value

 # Verifique se Num_CDA está vazio
    if Num_CDA is None:
        print("Num_CDA está vazio. Saindo do loop.")
        break  # Sai do loop se Num_CDA estiver vazio

    # Espere até que o elemento seja visível
    wait = WebDriverWait(driver, 10)
    element = wait.until(EC.visibility_of_element_located((By.XPATH, "//select[@name='tipo_id']//option[@value='3']")))

 # Botão Consultar
    btn_Selecionar = driver.find_element(By.XPATH, "//select[@name='tipo_id'] //option[@value='3']")
    btn_Selecionar.click()
    sleep(1)

 # Digitar a Identificação
    Campo_ID = driver.find_element(By.XPATH, "//input[@name='numero_id']")
    Campo_ID.clear()  # Limpar o campo antes de inserir um novo valor
    Campo_ID.send_keys(Num_ID)

 # Digitar o CDA
    Campo_CDA = driver.find_element(By.XPATH, "//input[@id='id_numero_daf']")
    Campo_CDA.clear()  # Limpar o campo antes de inserir um novo valor
    Campo_CDA.send_keys(Num_CDA)

     # Botão Consultar
    btn_Consultar = driver.find_element(By.XPATH, "//input[@type='button']")
    btn_Consultar.click()
    sleep(1)

# Verifique se a mensagem "Nenhum resultado com os critérios de consulta" está presente
    resultado_msg_elements = driver.find_elements(By.XPATH, "//p[@align='center']//b")
    resultado_msg = ""

    for element in resultado_msg_elements:
       resultado_msg += element.text
# Escreva a mensagem no Excel
       planilha.cell(row=row, column=column_index + 1, value=resultado_msg)

# Botão Voltar
    btn_Voltar = driver.find_element(By.XPATH, "//img[@id='ghost']")
    btn_Voltar.click()
    sleep(2)
    
# Tente encontrar o elemento até 10 segundos

    try:
        valor_total_element = driver.find_element(By.XPATH, "//td[contains(text(), 'Valor Total:')]/following-sibling::td/b/font[@color='#FF0000']")
        valor_total = valor_total_element.text
    except NoSuchElementException:
        # Elemento não encontrado, aguarde um pouco e tente novamente
        sleep(1)

# Encontre a próxima coluna disponível (vamos começar da coluna D)
    coluna_atual = column_index + 1

    # Escreva os valores nas colunas a partir da coluna D
    planilha.cell(row=row, column=coluna_atual, value=valor_total)
    coluna_atual += 1

 # Salve o arquivo Excel
    workbook.save(nome_arquivo_excel)

# Botão Voltar
    btn_Voltar = driver.find_element(By.XPATH, "//img[@id='ghost']")
    btn_Voltar.click()
    sleep(2)

    row += 1
    row2+= 1

# Fechar o WebDriver
driver.quit()
