from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.select import Select
import re
from selenium.common.exceptions import NoSuchElementException
from twocaptcha import TwoCaptcha
from anticaptchaofficial.recaptchav2proxyless import *
import pyautogui
from time import sleep
import openpyxl
from Chave import *

def new_func():
    pyautogui.press('backspace', presses=20)

def salvar_como_pdf(CDA):
    pyautogui.hotkey('ctrl', 'p')
    sleep(2)
    pyautogui.click(1519, 259)
    sleep(1)
    pyautogui.click(1477, 330)
    sleep(1)
    pyautogui.click(1493, 918)
    sleep(1)
    new_func()
    pyautogui.write(f'relatorio_{CDA}.pdf')
    pyautogui.click(716, 670)
    sleep(1)
    pyautogui.click(1033, 527)

# Inicializando o driver do Chrome
driver_service = Service(r'C:\Users\walter.oliveira\Documents\ProjetosPython\dev\chromedriver-win64\chromedriver.exe')
driver = webdriver.Chrome(service=driver_service)
link = 'https://www.dividaativa.pge.sp.gov.br/sc/pages/consultas/consultarDebito.jsf'
driver.get(link)

primeiro_registro2 = True  # Variável para controlar o primeiro registro

pyautogui.click(1480, 39)
sleep(2)
pyautogui.click(1868, 184)
sleep(2)

# Nome do arquivo Excel e nome da planilha
nome_arquivo_excel = r'C:\Users\walter.oliveira\Downloads\CDA_23_08_2024\São Paulo 6(Débitos IPVA SP - Pan Arre).xlsx'
nome_planilha_excel = 'Débitos IPVA SP - Pan Arre'

# Carregar a planilha Excel
workbook = openpyxl.load_workbook(nome_arquivo_excel)
planilha = workbook[nome_planilha_excel]

# Começando da linha 4, coluna 3 (C)
start_row = 331
column_index = 3

# Descobrir o número total de linhas na planilha
total_rows = planilha.max_row

# Loop para consultar todas as linhas da coluna 3, começando da linha 4
row = start_row  # Inicialize a variável row
primeiro_registro = True  # Variável para controlar o primeiro registro

while row <= total_rows:
    Num_CDA = planilha.cell(row=row, column=column_index).value

# Verifique se Num_CDA está vazio
    if Num_CDA is None:
        print("Num_CDA está vazio. Saindo do loop.")
        break  # Sai do loop se Num_CDA estiver vazio

   # Digitar o CDA
    Campo_CDA = driver.find_element(By.XPATH, "//input[@id='consultaDebitoForm:decTxtTipoConsulta:cdaEtiqueta']")
    Campo_CDA.clear()  # Limpar o campo antes de inserir um novo valor
    Campo_CDA.send_keys(Num_CDA)
   
    # Se for o primeiro registro, espere 35 segundos
    if primeiro_registro:
        chave_captcha = driver.find_element(By.CLASS_NAME, 'g-recaptcha').get_attribute('data-sitekey')

        solver = recaptchaV2Proxyless()
        solver.set_verbose(1)
        solver.set_key(chave_api)
        solver.set_website_url(link)
        solver.set_website_key(chave_captcha)

        resposta = solver.solve_and_return_solution()

        if resposta != 0:
            #print(resposta)
            # preencher o campo do token do captcha
            driver.execute_script(f"document.getElementById('g-recaptcha-response').innerHTML = '{resposta}'")
            driver.find_element(By.XPATH, "//*[@id='consultaDebitoForm:j_id78_body']/div[2]/input[2]").click()
        else:
            print(solver.err_string)

        primeiro_registro = False  # Defina como False após o primeiro registro

    # Botão Consultar
    btn_Consultar = driver.find_element(By.XPATH, "//*[@id='consultaDebitoForm:j_id78_body']/div[2]/input[2]")
    btn_Consultar.click()
    sleep(1)

    # Verifique se a mensagem "Nenhum resultado com os critérios de consulta" está presente
    resultado_msg_element = driver.find_elements(By.XPATH, "//div[@class='rich-panel-body ']//p")
    resultado_msg = resultado_msg_element[0].text if resultado_msg_element else ""

    # Escreva a mensagem no Excel
    planilha.cell(row=row, column=column_index + 1, value=resultado_msg)

    if "Nenhum resultado com os critérios de consulta" in resultado_msg:
     CDA = str(Num_CDA)
     salvar_como_pdf(CDA)
     row += 1  # Vá para a próxima linha e continue o loop
     continue
    # Consultar IPVA
    href_IPVA = driver.find_element(By.XPATH, "//a[@id='consultaDebitoForm:dataTable:0:lnkConsultaDebito']")
    href_IPVA.click()
    sleep(1)

    # Consultar Registro
    href_Registro = driver.find_element(By.XPATH, "//a[@href='#']")
    href_Registro.click()
    sleep(1) 
    
   # Verifique se o elemento valor_honorarios_element exis
    try:
      valor_honorarios_element = driver.find_element(By.XPATH, "//*[@id='consultaDebitoForm:j_id1130:4:j_id1138']")
      valor_honorarios = valor_honorarios_element.text
    except NoSuchElementException:
       valor_honorarios = 0
    try:
      valor_mora_multa_element = driver.find_element(By.XPATH,"//*[@id='consultaDebitoForm:j_id1130:5:j_id1138']")
      valor_mora_multa = valor_mora_multa_element.text
    except NoSuchElementException:
      valor_mora_multa = 0

    try:
      valor_juros_element = driver.find_element(By.XPATH,"//*[@id='consultaDebitoForm:j_id1130:2:j_id1138']")
      valor_juros = valor_juros_element.text
    except NoSuchElementException:
       valor_juros = 0

       # Extrair informações
    num_registro_element = driver.find_element(By.XPATH, "//*[@id='consultaDebitoForm:j_id1024']")
    num_registro = num_registro_element.find_element(By.TAG_NAME, "span").text
    
    numero_processo_element = driver.find_element(By.XPATH, "//*[@id='consultaDebitoForm:j_id1040']")
    numero_processo = numero_processo_element.find_element(By.TAG_NAME, "span").text
    
    numero_processo_outros_element = driver.find_element(By.XPATH, "//*[@id='consultaDebitoForm:j_id1049']")
    numero_processo_outros = numero_processo_outros_element.find_element(By.TAG_NAME, "span").text

    data_inscricao_element = driver.find_element(By.XPATH, "//*[@id='consultaDebitoForm:j_id1032']")
    data_inscricao = data_inscricao_element.find_element(By.TAG_NAME, "span").text

    situacao_element = driver.find_element(By.XPATH, "//*[@id='consultaDebitoForm:j_id1106']")
    situacao = situacao_element.find_element(By.TAG_NAME, "span").text
    
    saldo_element = driver.find_element(By.XPATH, "//*[@id='consultaDebitoForm:j_id1122']")
    saldo= saldo_element.find_element(By.TAG_NAME, "span").text
    
    valor_principal_element = driver.find_element(By.XPATH,"//*[@id='consultaDebitoForm:j_id1130:0:j_id1138']")
    valor_principal = valor_principal_element.text

    valor_multa_element = driver.find_element(By.XPATH,"//*[@id='consultaDebitoForm:j_id1130:3:j_id1138']")
    valor_multa = valor_multa_element.text
        
    placa_element = driver.find_element(By.XPATH,"//*[@id='consultaDebitoForm:detalheDebitoIpvaDeclarado:0:j_id1189']")
    placa = placa_element.text
    
    renavam_element = driver.find_element(By.XPATH,"//*[@id='consultaDebitoForm:detalheDebitoIpvaDeclarado:0:j_id1191']")
    renavam = renavam_element.text
    
    chassi_element = driver.find_element(By.XPATH,"//*[@id='consultaDebitoForm:detalheDebitoIpvaDeclarado:0:j_id1193']")
    chassi = chassi_element.text
   
    marca_modelo_element = driver.find_element(By.XPATH,"//*[@id='consultaDebitoForm:detalheDebitoIpvaDeclarado:0:j_id1195']")
    marca_modelo = marca_modelo_element.text
    
    ano_fab_element = driver.find_element(By.XPATH,"//*[@id='consultaDebitoForm:detalheDebitoIpvaDeclarado:0:j_id1197']")
    ano_fab = ano_fab_element.text
   
    ano_exercicio_element = driver.find_element(By.XPATH,"//*[@id='consultaDebitoForm:detalheDebitoIpvaDeclarado:0:j_id1199']")
    ano_exercicio = ano_exercicio_element.text
    
    dt_parcelas_element = driver.find_element(By.XPATH,"//*[@id='consultaDebitoForm:detalheDebitoIpvaDeclarado:0:j_id1201']")
    dt_parcelas = dt_parcelas_element.text
  
    
    # Encontre a próxima coluna disponível (vamos começar da coluna D)
    coluna_atual = column_index + 1

    # Escreva os valores nas colunas a partir da coluna D

    planilha.cell(row=row, column=coluna_atual, value=data_inscricao)
    coluna_atual += 1

    planilha.cell(row=row, column=coluna_atual, value=numero_processo)
    coluna_atual += 1

    planilha.cell(row=row, column=coluna_atual, value=numero_processo_outros)
    coluna_atual += 1

    planilha.cell(row=row, column=coluna_atual, value=situacao)
    coluna_atual += 1

    planilha.cell(row=row, column=coluna_atual, value=saldo)
    coluna_atual += 1

    planilha.cell(row=row, column=coluna_atual, value=valor_principal)
    coluna_atual += 1

    planilha.cell(row=row, column=coluna_atual, value=valor_juros)
    coluna_atual += 1

    planilha.cell(row=row, column=coluna_atual, value=valor_multa)
    coluna_atual += 1

    planilha.cell(row=row, column=coluna_atual, value=valor_honorarios)
    coluna_atual += 1

    planilha.cell(row=row, column=coluna_atual, value=valor_mora_multa)
    coluna_atual += 1

    planilha.cell(row=row, column=coluna_atual, value=placa)
    coluna_atual += 1

    planilha.cell(row=row, column=coluna_atual, value=renavam)
    coluna_atual += 1

    planilha.cell(row=row, column=coluna_atual, value=chassi)
    coluna_atual += 1

    planilha.cell(row=row, column=coluna_atual, value=marca_modelo)
    coluna_atual += 1

    planilha.cell(row=row, column=coluna_atual, value=ano_fab)
    coluna_atual += 1

    planilha.cell(row=row, column=coluna_atual, value=ano_exercicio)
    coluna_atual += 1

    planilha.cell(row=row, column=coluna_atual, value=dt_parcelas)
    
    # Salve o arquivo Excel
    workbook.save(nome_arquivo_excel)
    CDA = str(Num_CDA)
    salvar_como_pdf(CDA)

    # Botão Voltar
    wait = WebDriverWait(driver, 15)
    btn_Voltar = wait.until(EC.presence_of_element_located((By.XPATH, "//input[@id='consultaDebitoForm:btnVoltarDetalheDebito']")))
    #driver.find_element(By.XPATH, "//input[@name='consultaDebitoForm:btnVoltarDetalheDebito']")
    btn_Voltar.click()
    sleep(2)

     # Botão Voltar
    wait = WebDriverWait(driver, 15)
    btn_Voltar1 = wait.until(EC.presence_of_element_located((By.XPATH, "//*[@id='consultaDebitoForm:consultaDebito']/div[2]/input")))
      #driver.find_element(By.XPATH, "//input[@name='consultaDebitoForm:j_id264']")
    btn_Voltar1.click()
    sleep(2)
    row += 1

# Fechar o WebDriver
driver.quit()

