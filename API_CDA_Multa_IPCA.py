from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pyautogui
from time import sleep
import openpyxl

def new_func():
    pyautogui.press('backspace',presses=20)

# entrar no site da -https://www.dividaativa.pge.sp.gov.br/sc/pages/consultas/consultarDebito.jsf
chrome_driver_path = r'C:\ProjetorPython\dev\chromedriver-win64\chromedriver.exe'
driver = webdriver.Chrome(executable_path=chrome_driver_path)
driver.get('https://www.dividaativa.pge.sp.gov.br/sc/pages/consultas/consultarDebito.jsf')

pyautogui.click(982,25)

# Nome do arquivo Excel e nome da planilha
nome_arquivo_excel = 'C:\\ProjetorPython\\Modelo_Multa_IPCA.xlsx'
nome_planilha_excel = 'Débitos IPCA SP - Pan Arre'

# Carregar a planilha Excel
workbook = openpyxl.load_workbook(nome_arquivo_excel)
planilha = workbook[nome_planilha_excel]

# Começando da linha 4, coluna 3 (C)
start_row = 4
column_index = 3


# Descobrir o número total de linhas na planilha
total_rows = planilha.max_row

# Loop para consultar todas as linhas da coluna 3, começando da linha 4
row = start_row  # Inicialize a variável row
primeiro_registro = True  # Variável para controlar o primeiro registro

while row <= total_rows:
    Num_CDA = planilha.cell(row=row, column=column_index).value

# Verifique se Num_CDA está vazio
    if Num_CDA is None:
        print("Num_CDA está vazio. Saindo do loop.")
        break  # Sai do loop se Num_CDA estiver vazio

   # Digitar o CDA
    Campo_CDA = driver.find_element(By.XPATH, "//input[@id='consultaDebitoForm:decTxtTipoConsulta:cdaEtiqueta']")
    Campo_CDA.clear()  # Limpar o campo antes de inserir um novo valor
    Campo_CDA.send_keys(Num_CDA)

    # Se for o primeiro registro, espere 35 segundos
    if primeiro_registro:
        sleep(35)
        primeiro_registro = False  # Defina como False após o primeiro registro

    # Botão Consultar
    btn_Consultar = driver.find_element(By.XPATH, "//input[@name='consultaDebitoForm:j_id104']")
    btn_Consultar.click()
    sleep(1)

    # Clique no campo para Multa IPCA
    Campo_Multa_IPCA = driver.find_element(By.XPATH, "//a[@id='consultaDebitoForm:dataTable:0:lnkConsultaDebito']")
    Campo_Multa_IPCA.click()
    sleep(1)

   # Consulta Registro
    Campo_CDA = driver.find_element(By.XPATH, "//a[@href='#']")
    Campo_CDA.click()
    sleep(1) 

     # Verifique se a mensagem "Nenhum resultado com os critérios de consulta" está presente
    resultado_msg_element = driver.find_elements(By.XPATH, "//div[@class='rich-panel-body ']//p")
    resultado_msg = resultado_msg_element[0].text if resultado_msg_element else ""

   # Escreva a mensagem no Excel
    planilha.cell(row=row, column=column_index + 1, value=resultado_msg)

    if "Nenhum resultado com os critérios de consulta" in resultado_msg:
        CDA = str(num_registro)
        #Gerar PDF da Tela
        pyautogui.click(1348,51)
        sleep(2)
        pyautogui.click(1097,291)
        sleep(2)
        pyautogui.click(1147,163)
        sleep(2)
        pyautogui.click(1026,205)
        sleep(2)
        pyautogui.click(1045,672)
        sleep(2)
        pyautogui.click(251,376)
        sleep(2)
        new_func()
        sleep(2)
        pyautogui.write(CDA)
        sleep(2)
        pyautogui.click(510,443)
        sleep(2)
        row += 1  # Vá para a próxima linha e continue o loop
        continue
   
    # Extrair informações

    num_registro_element = driver.find_element(By.XPATH, "//div[@id='consultaDebitoForm:j_id1008']")
    num_registro = num_registro_element.find_element(By.TAG_NAME, "span").text
    
    numero_processo_element = driver.find_element(By.XPATH, "//div[@id='consultaDebitoForm:j_id1024']")
    numero_processo = numero_processo_element.find_element(By.TAG_NAME, "span").text
    
    numero_processo_outros_element = driver.find_element(By.XPATH, "//div[@id='consultaDebitoForm:j_id1033']")
    numero_processo_outros = numero_processo_outros_element.find_element(By.TAG_NAME, "span").text

    numero_Aiim_element = driver.find_element(By.XPATH, "//div[@id='consultaDebitoForm:j_id1098']")
    numero_Aiim= numero_Aiim_element.find_element(By.TAG_NAME, "span").text

    data_inscricao_element = driver.find_element(By.XPATH, "//div[@id='consultaDebitoForm:j_id1016']")
    data_inscricao = data_inscricao_element.find_element(By.TAG_NAME, "span").text

    situacao_element = driver.find_element(By.XPATH, "//div[@id='consultaDebitoForm:j_id1090']")
    situacao = situacao_element.find_element(By.TAG_NAME, "span").text
    
    saldo_element = driver.find_element(By.XPATH, "//div[@id='consultaDebitoForm:j_id1106']")
    saldo= saldo_element.find_element(By.TAG_NAME, "span").text
    
    valor_principal_element = driver.find_element(By.XPATH,"//tbody[@id='consultaDebitoForm:j_id1114:tb']//tr[contains(@class,'rich-table-row')]//td[@id='consultaDebitoForm:j_id1114:0:j_id1122']")
    valor_principal = valor_principal_element.text

    correcao_monetaria_element = driver.find_element(By.XPATH,"//tbody[@id='consultaDebitoForm:j_id1114:tb']//tr[contains(@class,'rich-table-row')]//td[@id='consultaDebitoForm:j_id1114:1:j_id1122']")
    correcao_monetaria = correcao_monetaria_element.text

    valor_mora_multa_element = driver.find_element(By.XPATH,"//tbody[@id='consultaDebitoForm:j_id1114:tb']//td[@id='consultaDebitoForm:j_id1114:2:j_id1122']")
    valor_mora_multa = valor_mora_multa_element.text

    valor_honorarios_element = driver.find_element(By.XPATH, "//tbody[@id='consultaDebitoForm:j_id1114:tb']//td[@id='consultaDebitoForm:j_id1114:3:j_id1122']")
    valor_honorarios = valor_honorarios_element.text

    data_element = driver.find_element(By.XPATH,"//tbody[@id='consultaDebitoForm:j_id1134:tb']//td[@id='consultaDebitoForm:j_id1134:0:j_id1149']")
    data = data_element.text
    
    valor_element = driver.find_element(By.XPATH,"//tbody[@id='consultaDebitoForm:j_id1134:tb']//td[@id='consultaDebitoForm:j_id1134:0:j_id1151']")
    valor= valor_element.text
    
    data_inicio_juros_element = driver.find_element(By.XPATH,"//tbody[@id='consultaDebitoForm:j_id1134:tb']//td[@id='consultaDebitoForm:j_id1134:0:j_id1154']")
    data_inicio_juros = data_inicio_juros_element.text
   
    data_inicio_correcao_element = driver.find_element(By.XPATH,"//tbody[@id='consultaDebitoForm:j_id1134:tb']//td[@id='consultaDebitoForm:j_id1134:0:j_id1156']")
    data_inicio_correcao = data_inicio_correcao_element.text
    
    # Encontre a próxima coluna disponível (vamos começar da coluna D)
    coluna_atual = column_index + 1
  

    # Escreva os valores nas colunas a partir da coluna D

    planilha.cell(row=row, column=coluna_atual, value=data_inscricao)
    coluna_atual += 1

    planilha.cell(row=row, column=coluna_atual, value=numero_processo)
    coluna_atual += 1

    planilha.cell(row=row, column=coluna_atual, value=numero_processo_outros)
    coluna_atual += 1

    planilha.cell(row=row, column=coluna_atual, value=situacao)
    coluna_atual += 1

    planilha.cell(row=row, column=coluna_atual, value=numero_Aiim)
    coluna_atual += 1

    planilha.cell(row=row, column=coluna_atual, value=saldo)
    coluna_atual += 1

    planilha.cell(row=row, column=coluna_atual, value=valor_principal)
    coluna_atual += 1

    planilha.cell(row=row, column=coluna_atual, value=correcao_monetaria)
    coluna_atual += 1

    planilha.cell(row=row, column=coluna_atual, value=valor_mora_multa)
    coluna_atual += 1

    planilha.cell(row=row, column=coluna_atual, value=valor_honorarios)
    coluna_atual += 1

    planilha.cell(row=row, column=coluna_atual, value=data)
    coluna_atual += 1

    planilha.cell(row=row, column=coluna_atual, value=valor)
    coluna_atual += 1

    planilha.cell(row=row, column=coluna_atual, value=data_inicio_juros)
    coluna_atual += 1

    planilha.cell(row=row, column=coluna_atual, value=data_inicio_correcao)
    coluna_atual += 1

    # Salve o arquivo Excel
    workbook.save(nome_arquivo_excel)
    CDA = str(num_registro)
   #Gerar PDF da Tela
    pyautogui.click(1348,51)
    sleep(2)
    pyautogui.click(1097,291)
    sleep(2)
    pyautogui.click(1147,163)
    sleep(2)
    pyautogui.click(1026,205)
    sleep(2)
    pyautogui.click(1045,672)
    sleep(2)
    pyautogui.click(247,377)
    sleep(2)
    new_func()
    sleep(2)
    pyautogui.write(CDA)
    sleep(2)
    pyautogui.click(514,443)
    sleep(2)

     # Botão Voltar
    btn_Voltar = driver.find_element(By.XPATH, "//input[@name='consultaDebitoForm:btnVoltarDetalheDebito']")
    btn_Voltar.click()
    sleep(5)

     # Botão Voltar
    wait = WebDriverWait(driver, 35) # Espere até 20 segundos
    btn_Voltar1 = wait.until(EC.presence_of_element_located((By.XPATH, "//input[@name='consultaDebitoForm:j_id264']")))
    btn_Voltar1.click()
    sleep(2)
    row += 1

# Fechar o WebDriver
driver.quit()

