from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.select import Select
import re
from selenium.common.exceptions import NoSuchElementException
import pyautogui
from time import sleep
import openpyxl
from selenium.common.exceptions import TimeoutException

def apagar():
    pyautogui.press('backspace', presses=20)

def salvar_como_pdf(CDA):
    pyautogui.hotkey('ctrl', 'p')
    sleep(2)
    pyautogui.click(1519, 259)
    sleep(1)
    pyautogui.click(1477, 330)
    sleep(1)
    pyautogui.click(1493, 918)
    sleep(1)
    apagar()
    pyautogui.write(f'relatorio_BV_{CDA}.pdf')
    pyautogui.click(716, 670)
    sleep(1)
    pyautogui.click(1033, 527)
  
# Inicializando o driver do Chrome
driver_service = Service(r'C:\Users\walter.oliveira\Documents\ProjetosPython\dev\chromedriver-win64\chromedriver.exe')
driver = webdriver.Chrome(service=driver_service)
link = 'http://receitaonline.fazenda.mg.gov.br/rol/dae/'
driver.get(link)

pyautogui.click(1480, 39)
sleep(2)
pyautogui.click(1868, 184)
sleep(2)

#printlongo()

# Nome do arquivo Excel e nome da planilha
nome_arquivo_excel = r'C:\Users\walter.oliveira\Documents\ProjetosPython\dev\Bichara_Dev\repository\Atualização CDAs - BV - 5316779-54.2023.8.13.0024 - ref. 08_2024.xlsx'
nome_planilha_excel = '5316779-54.2023.8.13.0024'

# Carregar a planilha Excel
workbook = openpyxl.load_workbook(nome_arquivo_excel)
planilha = workbook[nome_planilha_excel]

# Começando da linha 4, coluna 3 (C)
Linha = 143
Coluna_2 = 2
coluna_3 = 3

# Descobrir o número total de linhas na planilha
total_rows = planilha.max_row

# Loop para consultar todas as linhas da coluna 3, começando da linha 4
linha = Linha  # Inicialize a variável row
primeiro_registro = True  # Variável para controlar o primeiro registro

# Localizar o seletor pelo caminho especificado
select_element = WebDriverWait(driver, 10).until(
    EC.element_to_be_clickable(
        (By.XPATH, "/html/body/div/div/div[3]/div[1]/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td/form/table[1]/tbody/tr[1]/td[2]/select"))
    )
 
# Criar um objeto Select
select = Select(select_element)
 
# Selecionar o exercício na lista suspensa
try:
    select.select_by_visible_text('CNPJ')
except Exception as e:
    print(f"Erro ao selecionar o exercício: {e}")
    sleep(2)

while linha <= total_rows:
    
    # Localizar o seletor pelo caminho especificado
    select_element = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable(
            (By.XPATH, "/html/body/div/div/div[3]/div[1]/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td/form/table[1]/tbody/tr[1]/td[2]/select"))
        )
 
    # Criar um objeto Select
    select = Select(select_element)
    
    # Selecionar o exercício na lista suspensa
    try:
        select.select_by_visible_text('CNPJ')
    except Exception as e:
        print(f"Erro ao selecionar o exercício: {e}")
        sleep(2)
    
    Num_CDA = planilha.cell(row=linha, column=coluna_3).value
    Num_CNPJ = planilha.cell(row=linha, column=Coluna_2).value
    
    if Num_CDA is None:
        print("Num_CDA está vazio. Saindo do loop.")
        break  # Sai do loop se Num_CDA estiver vazio
    
    Campo_CNPJ = driver.find_element(By.XPATH, "/html/body/div/div/div[3]/div[1]/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td/form/table[1]/tbody/tr[2]/td[2]/input")
    Campo_CNPJ.clear()  # Limpar o campo antes de inserir um novo valor
    Campo_CNPJ.send_keys(Num_CNPJ)
    
    Campo_CDA = driver.find_element(By.XPATH, "//*[@id='id_numero_daf']")
    Campo_CDA.clear()  # Limpar o campo antes de inserir um novo valor
    Campo_CDA.send_keys(Num_CDA)
    
    # Botão Consultar
    btn_Consultar = driver.find_element(By.XPATH, "/html/body/div/div/div[3]/div[1]/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td/form/table[2]/tbody/tr/td/input[1]")
    btn_Consultar.click()
    sleep(1)
    
    try:
        Valor_Total_Elemento = driver.find_element(By.XPATH, "/html/body/div/div/div[3]/div[1]/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr[9]/td[2]/b/font")
        Valor_Total = Valor_Total_Elemento.text

        coluna_atual = coluna_3 + 1

        planilha.cell(row=linha, column=coluna_atual, value=Valor_Total)
        coluna_atual += 1
        
        # Salve o arquivo Excel
        workbook.save(nome_arquivo_excel)
        CDA = str(Num_CDA)
        salvar_como_pdf(CDA)    

        linha += 1
        
        sleep(1.5)
        # Botão Voltar
        pyautogui.click(1498,927)
        sleep(1)
        
    except:        
        Mensagem_Erro = "Mensagem Broker: QUITADO NAO PERMITE PAGAMENTO. DUVIDA, CONTATE ADM.FAZENDARIA"
        
        coluna_atual = coluna_3 + 1
        
        planilha.cell(row=linha, column=coluna_atual, value=Mensagem_Erro)
        
        coluna_atual += 1
        
        workbook.save(nome_arquivo_excel)
        CDA = str(Num_CDA)
        salvar_como_pdf(CDA)    

        linha += 1
        
        sleep(1.5)
        # Botão Voltar
        pyautogui.click(1513,739)
        sleep(1)

# Fechar o WebDriver
driver.quit()

    
    
    

