from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pyautogui
from time import sleep
import openpyxl
import os
import win32com.client
import email
import re
def new_func():
    pyautogui.press('backspace',presses=20)

# entrar no site da -hhttps://projuris.bancovotorantim.com.br/projuris
chrome_driver_path = r'C:\ProjetorPython\Consultar_CDA\Consultar_CDA\chromedriver-win64\chromedriver.exe'
driver = webdriver.Chrome(executable_path=chrome_driver_path)
driver.get('https://bv.projuris.com.br/projuris')

pyautogui.click(982,25)

# Nome do arquivo Excel e nome da planilha
nome_arquivo_excel = r'C:\ProjetorPython\Análise EFs(7).xlsx'
nome_planilha_excel = 'Débitos IPVA SP - Pan Arre'

# Carregar a planilha Excel
workbook = openpyxl.load_workbook(nome_arquivo_excel)
planilha = workbook[nome_planilha_excel]

# Começando da linha 4, coluna 3 (C)
start_row = 4
column_index = 3

# Descobrir o número total de linhas na planilha
total_rows = planilha.max_row

# Loop para consultar todas as linhas da coluna 3, começando da linha 4
row = start_row  # Inicialize a variável row
primeiro_registro = True  # Variável para controlar o primeiro registro

Projuris_USER = os.getenv('Projuris_USER') 
Projuris_PWD = os.getenv('Projuris_PWD')

if Projuris_USER is None or Projuris_PWD is None:
    print("As variáveis de ambiente Projuris_USER e/ou Projuris_PWD não estão definidas corretamente.")
else:
  # Digitar o Login e Senha

    Campo_Login = driver.find_element(By.XPATH, " //input[@id = 'LOGINprojuris/LoginVO_*_login']")
    Campo_Login.clear()  # Limpar o campo antes de inserir um novo valor
    Campo_Login.send_keys(Projuris_USER)

    Campo_Senha = driver.find_element(By.XPATH,"//input[@id = 'SENHAprojuris/LoginVO_*_login']")
    Campo_Senha.clear()
    Campo_Senha.send_keys(Projuris_PWD)

    pyautogui.click(724,490) 

    # Abra o Outlook
    outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")

    # Pasta de entrada (Inbox)
    inbox = outlook.GetDefaultFolder(6)  # O valor 6 representa a pasta de entrada (Inbox)

    # Aguarde algum tempo para garantir que o Outlook seja aberto
    sleep(50)

    # Obtenha a última mensagem
    messages = inbox.Items
    message = messages.GetLast()

    # Extraia o texto da mensagem
    body = message.Body

  # Use expressões regulares para encontrar o código de verificação
    verification_code = re.search(r'seu código de autenticação para acesso ao Projuris é: (\d+)', body)
    fator = verification_code.group(1)
    
  # Feche o Outlook
    message.Close(0)

    Campo_Fator_Atenticacao = driver.find_element(By.XPATH,"//input[@id='CODIGO_AUTENTICACAOprojuris/LoginVO_*_duplo_fator_autenticacao']")
    Campo_Fator_Atenticacao.clear()
    Campo_Fator_Atenticacao.send_keys(str(fator))

    pyautogui.click(800,550)
    sleep(2)

    pyautogui.click(1081,660)
    sleep(2)

    while row <= total_rows:
       Num_CDA = planilha.cell(row=row, column=column_index).value
# Verifique se Num_CDA está vazio
       if Num_CDA is None:
        print("Num_CDA está vazio. Saindo do loop.")

       break  # Sai do loop se Num_CDA estiver vazio
 # Botão Processo   
    sleep(2)
    btn_Processo = driver.find_element(By.XPATH,"//div [@tree-node-id='PR']//a [@class='x-tree-node-anchor']") 
    btn_Processo.click()
    sleep(2)
# Botão Abrir Documento
    btn_AbrirDoc = driver.find_element(By.XPATH,"//div [@tree-node-id='PR_lin_862569224']//a [@class='x-tree-node-anchor']")
    btn_AbrirDoc.click()
    sleep(2)
    pyautogui.click(488,165)
    sleep(2)
    pyautogui.write(str(Num_CDA))
    sleep(2)
    pyautogui.click(762,217)
    sleep(2)
  
