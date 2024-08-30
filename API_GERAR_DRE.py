from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.select import Select
import re
import os
from twocaptcha import TwoCaptcha
from anticaptchaofficial.recaptchav2proxyless import *
import pyautogui
from time import sleep
import openpyxl
from Chave import *
from bs4 import BeautifulSoup

def renomear_arquivo_downloads(cda, num_processo):
    # Diretório de downloads
    pasta_downloads = os.path.expanduser('~/Downloads')
    
    # Caminho completo do arquivo original
    arquivo_original = os.path.join(pasta_downloads, 'DARE.pdf')
    
    # Verifica se o arquivo existe
    if not os.path.isfile(arquivo_original):
        print(f'O arquivo {arquivo_original} não foi encontrado.')
        return
    
    # Novo nome do arquivo
    novo_nome = f'CDA_{cda}_{num_processo}.pdf'
    novo_caminho = os.path.join(pasta_downloads, novo_nome)
    
    # Renomear o arquivo
    os.rename(arquivo_original, novo_caminho)
    print(f'Arquivo renomeado para {novo_caminho}')

def extrair_informacoes(CDA, num_processo):
     
    # Extrair informações
    parte_reu_element = driver.find_element(By.XPATH, "//*[@id='consultaDebitoForm:j_id389']")
    parte_reu = parte_reu_element.find_element(By.TAG_NAME, "span").text

    parte_autor_element = driver.find_element(By.XPATH, "//*[@id='consultaDebitoForm:j_id381']")
    parte_antor = parte_autor_element.find_element(By.TAG_NAME, "span").text

    vara_element = driver.find_element(By.XPATH, "//*[@id='consultaDebitoForm:j_id397']")
    vara = vara_element.find_element(By.TAG_NAME, "span").text

    comarca_foro_element = driver.find_element(By.XPATH, "//*[@id='consultaDebitoForm:j_id405']")
    comarca_foro = comarca_foro_element.find_element(By.TAG_NAME, "span").text

    data_inscricao_element = driver.find_element(By.XPATH, "//*[@id='consultaDebitoForm:j_id332']")
    data_inscricao = data_inscricao_element.find_element(By.TAG_NAME, "span").text

    situacao_element = driver.find_element(By.XPATH, "//*[@id='consultaDebitoForm:j_id356']")
    situacao = situacao_element.find_element(By.TAG_NAME, "span").text
    
    saldo_element = driver.find_element(By.XPATH, "//span[@id='consultaDebitoForm:gerarGare']")
    saldo = saldo_element.find_element(By.TAG_NAME, "div").text
    saldo_devedor = saldo.strip('Saldo Devedor (R$):')  # Remove Saldo Devedor (R$) do início 
 
    valor_principal_element = driver.find_element(By.XPATH,"//*[@id='consultaDebitoForm:j_id420:0:j_id430'] ")
    valor_principal = valor_principal_element.text

    valor_juros_element = driver.find_element(By.XPATH,"//*[@id='consultaDebitoForm:j_id420:2:j_id430']")
    valor_juros = valor_juros_element.text

    valor_honorarios_element = driver.find_element(By.XPATH, "//*[@id='consultaDebitoForm:j_id420:3:j_id430']")
    valor_honorarios = valor_honorarios_element.text

     # Encontre a próxima coluna disponível (vamos começar da coluna D)
    coluna_atual = column_index + 1

    # Escreva os valores nas colunas a partir da coluna D
    planilha.cell(row=row, column=coluna_atual, value=numero_processo)
    coluna_atual += 1

    planilha.cell(row=row, column=coluna_atual, value=parte_antor)
    coluna_atual += 1

    planilha.cell(row=row, column=coluna_atual, value=parte_reu)
    coluna_atual += 1

    planilha.cell(row=row, column=coluna_atual, value=situacao)
    coluna_atual += 1
     
    planilha.cell(row=row, column=coluna_atual, value=data_inscricao)
    coluna_atual += 1

    planilha.cell(row=row, column=coluna_atual, value=vara)
    coluna_atual += 1

    planilha.cell(row=row, column=coluna_atual, value=comarca_foro)
    coluna_atual += 1

    planilha.cell(row=row, column=coluna_atual, value=valor_principal)
    coluna_atual += 1

    planilha.cell(row=row, column=coluna_atual, value=valor_juros)
    coluna_atual += 1

    planilha.cell(row=row, column=coluna_atual, value=valor_honorarios)
    coluna_atual += 1 

    planilha.cell(row=row, column=coluna_atual, value=saldo_devedor)
    coluna_atual += 1

    planilha.cell(row=row, column=coluna_atual, value=(f'relatorio_{CDA}_{num_processo}.pdf'))
    coluna_atual += 1

# Inicializando o driver do Chrome
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
link = 'https://www.dividaativa.pge.sp.gov.br/sc/pages/consultas/consultarDebito.jsf'
driver.get(link)

pyautogui.click(1474,44)
sleep(3)
pyautogui.click(1870,180)
sleep(3)

# Nome do arquivo Excel e nome da planilha
nome_arquivo_excel = r'C:\Users\walter.oliveira\Documents\ProjetosPython\dev\Bichara_Dev\repository\Template_Ipva e Taxa_Judiciaria - VDG - 29.08.xlsx'
nome_planilha_excel = 'Débitos Taxa Judiciaria'

# Carregar a planilha Excel
workbook = openpyxl.load_workbook(nome_arquivo_excel)
planilha = workbook[nome_planilha_excel]

# Começando da linha 4, coluna 3 (C)
start_row = 4
column_index = 3

# Descobrir o número total de linhas na planilha
total_rows = planilha.max_row

# Loop para consultar todas as linhas da coluna 3, começando da linha 4
row = start_row  # Inicialize a variável row
primeiro_registro = True  # Variável para controlar o primeiro registro

solver = TwoCaptcha(Var_2captcha)

while row <= total_rows:
    Num_CDA = planilha.cell(row=row, column=column_index).value

    # Verifique se Num_CDA está vazio
    if Num_CDA is None:
        print("Num_CDA está vazio. Saindo do loop.")
        break  # Sai do loop se Num_CDA estiver vazio   
    
    # Digitar o CDA
    Campo_CDA = driver.find_element(By.XPATH, "//input[@id='consultaDebitoForm:decTxtTipoConsulta:cdaEtiqueta']")
    Campo_CDA.clear()  # Limpar o campo antes de inserir um novo valor
    Campo_CDA.send_keys(Num_CDA)
   
    # Se for o primeiro registro, espere 35 segundos
    if primeiro_registro:
        chave_captcha = driver.find_element(By.CLASS_NAME, 'g-recaptcha').get_attribute('data-sitekey')

        solver = recaptchaV2Proxyless()
        solver.set_verbose(1)
        solver.set_key(chave_api)
        solver.set_website_url(link)
        solver.set_website_key(chave_captcha)

        resposta = solver.solve_and_return_solution()

        if resposta != 0:
           # print(resposta)
            # preencher o campo do token do captcha
            driver.execute_script(f"document.getElementById('g-recaptcha-response').innerHTML = '{resposta}'")
            driver.find_element(By.XPATH, "//*[@id='consultaDebitoForm:j_id78_body']/div[2]/input[2]").click()
        else:
            print(solver.err_string)

        primeiro_registro = False  # Defina como False após o primeiro registro
        
    # Botão Consultar
    btn_Consultar = driver.find_element(By.XPATH, "//*[@id='consultaDebitoForm:j_id78_body']/div[2]/input[2]")
    btn_Consultar.click()
    sleep(2)   
    # Verifique se a mensagem "Nenhum resultado com os critérios de consulta" está presente
    resultado_msg_element = driver.find_elements(By.XPATH, "//div[@class='rich-panel-body ']//p")
    resultado_msg = resultado_msg_element[0].text if resultado_msg_element else ""

    # Escreva a mensagem no Excel
    planilha.cell(row=row, column=column_index + 1, value=resultado_msg)
    if "Nenhum resultado com os critérios de consulta" in resultado_msg:
        CDA = str(Num_CDA)
        num_processo = '0'
        # Gerar PDF da Tela
        salvar_como_pdf(CDA, num_processo)
        row += 1  # Vá para a próxima linha e continue o loop 
        continue

    # Consultar IPVA
    href_IPVA = driver.find_element(By.XPATH, "//*[@id='consultaDebitoForm:dataTable:0:lnkConsultaDebito']")
    href_IPVA.click()
    sleep(2)

    btn_liquidar = driver.find_element(By.XPATH, "//*[@id='consultaDebitoForm:dataTable2:0:lnkLiquidarDebito']" )
    btn_liquidar.click()
    sleep(2)

    numero_processo_element = driver.find_element(By.XPATH, "//*[@id='consultaDebitoForm:j_id373']")
    numero_processo = numero_processo_element.find_element(By.TAG_NAME, "span").text

    extrair_informacoes(Num_CDA, numero_processo)
    sleep(2)

    # Verificar se o elemento está visível e clicável
    btn_Gerar_Gare = WebDriverWait(driver, 15).until(
        EC.visibility_of_element_located((By.XPATH, "//*[@id='consultaDebitoForm:btnDownload']"))
    )
    btn_Gerar_Gare = WebDriverWait(driver, 15).until(
        EC.element_to_be_clickable((By.XPATH, "//*[@id='consultaDebitoForm:btnDownload']"))
    )
    btn_Gerar_Gare.click()
    sleep(2)

    pyautogui.click(1348,450)
    sleep(2)
    # Esperar até que o painel modal desapareça
    WebDriverWait(driver, 15).until(
        EC.invisibility_of_element((By.ID, "modalPanelMsgGerarGareDiv"))
    )

    btn_download = WebDriverWait(driver, 15).until(
        EC.element_to_be_clickable((By.XPATH, "//*[@id='consultaDebitoForm:btnDownloadGare']"))
    )
    btn_download.click()
    sleep(5)

    renomear_arquivo_downloads(Num_CDA, numero_processo) 
    sleep(3)

    # Botão Voltar
    btn_Voltar = driver.find_element(By.XPATH, "//*[@id='consultaDebitoForm:j_id858_body']/div[1]/input[1]")
    btn_Voltar.click()
    sleep(3)
    
    # Botão Voltar
    btn_Voltar1 = driver.find_element(By.XPATH, "//*[@id='consultaDebitoForm:btnVoltar']")
    btn_Voltar1.click()
    sleep(3)

    # Botão Voltar
    btn_Voltar2 = driver.find_element(By.XPATH, "//input[@name='consultaDebitoForm:j_id284']")
    btn_Voltar2.click()
    sleep(3)
    row += 1
# Salve o arquivo Excel atualizado
    workbook.save(nome_arquivo_excel)
    continue  
# Fechar o WebDriver
driver.quit()