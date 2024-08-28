import openpyxl
from tkinter import Tk
from tkinter.filedialog import askopenfilename

def separar_areas_interesse(texto):
    dados = [item.strip() for item in texto.replace("[", "").replace("]", "").replace('"', "").split(";")]
    return [area for area in dados if area]

def separar_dados_pessoa_fisica(texto):
    dados = [item.strip() for item in texto.replace("[", "").replace("]", "").replace('"', "").split(",")]
    if len(dados) < 6:
        return ["", "", "", ""]
    nome = dados[0] if len(dados) > 0 else ""
    cpf = dados[1] if len(dados) > 1 else ""
    estado = dados[-2] if len(dados) > 4 else ""
    pais = dados[-1] if len(dados) > 5 else ""
    return [nome, cpf, estado, pais]

def separar_dados_estrangeira(texto):
    dados = [item.strip() for item in texto.replace("[", "").replace("]", "").replace('"', "").split(",")]
    if len(dados) < 3:
        return ["", "", ""]
    nome = dados[0] if len(dados) > 0 else ""
    estado = dados[-2] if len(dados) > 1 else ""
    pais = dados[-1] if len(dados) > 2 else ""
    return [nome, estado, pais]

def separar_dados_relacionamento(texto):
    # Remover colchetes externos e dividir os blocos
    contatos_brutos = texto.replace("[[", "").replace("]]", "").split("],[")

    contatos_separados = []
    
    for contato in contatos_brutos:
        # Remover colchetes e aspas extras, e dividir por vírgula
        dados = [item.strip().replace('"', '') for item in contato.replace("[", "").replace("]", "").split(",")]
        # Garantir que todos os contatos tenham pelo menos 4 elementos
        if len(dados) < 4:
            dados.extend([""] * (4 - len(dados)))
        contatos_separados.append(dados[:4])
    
    return contatos_separados

def separar_dados():
    Tk().withdraw()
    
    caminho_arquivo = askopenfilename(
        title="Selecione a planilha a ser processada",
        filetypes=[("Arquivos do Excel", "*.xls *.xlsx")]
    )
    
    if not caminho_arquivo:
        print("Nenhum arquivo foi selecionado.")
        return
    
    wb_origem = openpyxl.load_workbook(caminho_arquivo)
    ws_origem = wb_origem.active
    
    ws_destino = wb_origem.create_sheet(title="Dados Separados")
    
    cabecalho = [
        "CNPJ/CPF", "Nome Empresa", "Serviço", 
        "RequestId", "WorkflowStartedTimestamp", "WorkflowStartedTimestamp", "WorkflowStatus", 
        "Escritório Faturamento", "Tipo Cadastro", "Responsável", "Tipo Pessoa", 
        "Caso Pro Bono?", "Advogado Responsável", "Sócio Responsável",
        "Estado", "País"
    ]
    
    max_areas = 0
    max_contatos = 0
    
    for i in range(2, ws_origem.max_row + 1):
        texto_areas = ws_origem.cell(row=i, column=42).value
        texto_ao = ws_origem.cell(row=i, column=41).value
        
        if texto_areas:
            areas = separar_areas_interesse(texto_areas)
            max_areas = max(max_areas, len(areas))
        
        if texto_ao:
            contatos = separar_dados_relacionamento(texto_ao)
            max_contatos = max(max_contatos, len(contatos))
    
    for j in range(1, max_contatos + 1):
        cabecalho.extend([
            f"Contato_{j}", f"E-mail_{j}", f"Cargo_{j}", f"Tel_{j}"
        ])

    for j in range(1, max_areas + 1):
        cabecalho.append(f"Áreas de interesse_{j}")
    
    ws_destino.append(cabecalho)
    
    for i in range(2, ws_origem.max_row + 1):
        texto = ws_origem.cell(row=i, column=17).value
        texto2 = ws_origem.cell(row=i, column=18).value
        tipo_pessoa = ws_origem.cell(row=i, column=16).value
        texto_s = ws_origem.cell(row=i, column=19).value
        texto_ao = ws_origem.cell(row=i, column=41).value
        texto_areas = ws_origem.cell(row=i, column=42).value
        
        col_a = ws_origem.cell(row=i, column=1).value
        col_d = ws_origem.cell(row=i, column=4).value
        col_e = ws_origem.cell(row=i, column=5).value
        col_g = ws_origem.cell(row=i, column=7).value
        col_i = ws_origem.cell(row=i, column=9).value
        col_m = ws_origem.cell(row=i, column=13).value
        col_o = ws_origem.cell(row=i, column=15).value
        col_p = ws_origem.cell(row=i, column=16).value
        col_v = ws_origem.cell(row=i, column=22).value
        col_x = ws_origem.cell(row=i, column=24).value
        col_y = ws_origem.cell(row=i, column=25).value
        
        contatos_relacionamento = separar_dados_relacionamento(texto_ao) if texto_ao else [["", "", "", ""]]
        
        areas_interesse = separar_areas_interesse(texto_areas) if texto_areas else []
        
        if tipo_pessoa == "Pessoa Física" and texto2:
            dados_pessoa_fisica = separar_dados_pessoa_fisica(texto2)
            linha_destino = [
                dados_pessoa_fisica[1], dados_pessoa_fisica[0], "",  # CPF e nome nos mesmos campos de CNPJ e Nome Empresa
                col_a, col_d, col_e, col_g, 
                col_i, col_m, col_o, col_p, 
                col_v, col_x, col_y,
                dados_pessoa_fisica[2], dados_pessoa_fisica[3]
            ]
        elif tipo_pessoa == "Estrangeira" and texto_s:
            dados_estrangeira = separar_dados_estrangeira(texto_s)
            linha_destino = [
                "", dados_estrangeira[0], "",  # Nome na coluna "Nome Empresa"
                col_a, col_d, col_e, col_g, 
                col_i, col_m, col_o, col_p, 
                col_v, col_x, col_y,
                dados_estrangeira[1], dados_estrangeira[2]
            ]
        else:
            if texto:
                dados = processar_dados(texto)
                linha_destino = dados + [
                    col_a, col_d, col_e, col_g, 
                    col_i, col_m, col_o, col_p, 
                    col_v, col_x, col_y,
                    "", ""
                ]
        
        for contato in contatos_relacionamento:
            linha_destino.extend(contato)
        linha_destino.extend(areas_interesse + [""] * (max_areas - len(areas_interesse)))
        
        ws_destino.append(linha_destino)
    
    wb_origem.save(caminho_arquivo)
    print("Processamento concluído e arquivo salvo.")

def processar_dados(texto):
    # Remover colchetes e aspas duplas, e dividir os dados por vírgula
    dados = [item.strip() for item in texto.replace("[", "").replace("]", "").replace('"', "").split(",")]

    # Definir os valores padrão
    cnpj_cpf = dados[0] if len(dados) > 0 else ""
    nome_empresa = dados[1] if len(dados) > 1 else ""
    
    # Inicializar a variável do serviço
    servico = ""
    
    # Verificar a partir do terceiro item (índice 2) se existe um serviço
    for item in dados[2:]:
        item = item.strip()
        # Ignorar partes que são valores percentuais ou numéricos
        if "%" in item or item.replace('.', '', 1).isdigit():
            continue
        # Adicionar o item ao serviço se for relevante
        if servico == "":
            servico = item
        else:
            # Verificar se a próxima parte é um possível CNPJ/CPF ou nome e parar a concatenação
            if len(item) > 0 and (item[0].isdigit() or len(item) > 14):
                break
            servico += ", " + item

    return [cnpj_cpf, nome_empresa, servico]

separar_dados()
