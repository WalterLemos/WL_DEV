import win32com.client as win32
import openpyxl
from tqdm import tqdm

# Nome do arquivo Excel e nome da planilha
nome_arquivo_excel = 'C:\\temp\\Teste.xlsx'
nome_planilha_excel = 'Planilha1'

# Carregar a planilha Excel
workbook = openpyxl.load_workbook(nome_arquivo_excel)
planilha = workbook[nome_planilha_excel]

# Iniciar o Outlook
outlook = win32.Dispatch("Outlook.Application").GetNamespace("MAPI")

# Criar barra de progresso
total_rows = planilha.max_row   # Descontar o cabeçalho
progress_bar = tqdm(total=total_rows, desc='Pesquisando e-mails')

# Destinatário fixo (domínio)
destinatario_pesquisa = "gupy.com.br"

# Lista para armazenar as mensagens coletadas
mensagens_coletadas = []

# Percorrer todas as pastas do Outlook
for folder in outlook.Folders:
    # Ler remetente da planilha e realizar a pesquisa na pasta atual
    for row in planilha.iter_rows(min_row=1, values_only=True):  # Começar a partir da linha 2
        remetente_pesquisa = row[1]  # Remetente na primeira coluna

        # Realizar a pesquisa com base no remetente e destinatário
        search_folder = folder
        filter_query = "([SenderName] = '{}') AND ([To] = '*@{}')".format(remetente_pesquisa, destinatario_pesquisa)
        search_results = search_folder.Items.Restrict(filter_query)

        # Armazenar as mensagens coletadas na lista
        for email in search_results:
            mensagem = {
                'Pasta': folder.Name,
                'Remetente': remetente_pesquisa,
                'Destinatario': destinatario_pesquisa,
                'Assunto': email.Subject,
                'RemetenteEmail': email.SenderEmailAddress,
                'Data': email.ReceivedTime
            }
            mensagens_coletadas.append(mensagem)

        # Atualizar a barra de progresso
        progress_bar.update()

workbook.close()
progress_bar.close()  # Fechar a barra de progresso ao final

# Imprimir as mensagens coletadas
for mensagem in mensagens_coletadas:
    print(f'Pasta: {mensagem["Pasta"]}')
    print(f'Remetente da pesquisa: {mensagem["Remetente"]}')
    print(f'Destinatário da pesquisa: {mensagem["Destinatario"]}')
    print(f'Assunto: {mensagem["Assunto"]}')
    print(f'Remetente Email: {mensagem["RemetenteEmail"]}')
    print(f'Data: {mensagem["Data"]}')
    print('-' * 40)


