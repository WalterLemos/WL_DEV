import os
import tkinter as tk
from tkinter import filedialog, messagebox
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.select import Select
import re
import openpyxl
import threading
import time
import base64

# Configurações do Chrome para salvar como PDF
def configurar_chrome_options():
    chrome_options = Options()
    chrome_options.add_argument('--kiosk-printing')  # Ativa a impressão automática sem diálogo
    return chrome_options

def gerar_pdf_dinamico(driver, caminho_diretorio, nome_pdf):
    # Define o caminho completo do arquivo PDF usando o nome_pdf fornecido
    caminho_arquivo_pdf = os.path.join(caminho_diretorio, f"{nome_pdf}.pdf")

    # Usa a API DevTools para gerar o PDF da página atual
    result = driver.execute_cdp_cmd("Page.printToPDF", {
        "landscape": False,                 # Define orientação do PDF como retrato
        "displayHeaderFooter": False,       # Não exibe cabeçalho e rodapé
        "printBackground": True,            # Inclui o fundo da página no PDF
        "preferCSSPageSize": True           # Usa o tamanho da página definido por CSS
    })

    # Decodifica o PDF gerado (base64) e salva em um arquivo
    with open(caminho_arquivo_pdf, "wb") as file:
        file.write(base64.b64decode(result['data']))

    print(f"PDF gerado e salvo em: {caminho_arquivo_pdf}")

def start_process(excel_file, sheet_name, start_row, caminho_diretorio):
    # Caminho fixo para o ChromeDriver
    driver_path = r'C:\Users\walter.oliveira\Documents\ProjetosPython\dev\chromedriver-win64\chromedriver.exe'  # Substitua com o caminho real do seu chromedriver

    # Verifica se o diretório existe e é gravável
    if not os.path.exists(caminho_diretorio):
        print(f"Erro: O diretório {caminho_diretorio} não existe.")
        return
    elif not os.access(caminho_diretorio, os.W_OK):
        print(f"Erro: O diretório {caminho_diretorio} não é gravável.")
        return

    # Inicializando o driver do Chrome
    chrome_options = configurar_chrome_options()
    driver_service = Service(driver_path)
    driver = webdriver.Chrome(service=driver_service, options=chrome_options)
    link = 'http://receitaonline.fazenda.mg.gov.br/rol/dae/'
    driver.get(link)

    workbook = openpyxl.load_workbook(excel_file)
    planilha = workbook[sheet_name]

    total_rows = planilha.max_row
    linha = int(start_row)

    select_element = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable(
            (By.XPATH, "/html/body/div/div/div[3]/div[1]/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td/form/table[1]/tbody/tr[1]/td[2]/select"))
        )
    select = Select(select_element)
    
    try:
        select.select_by_visible_text('CNPJ')
    except Exception as e:
        print(f"Erro ao selecionar o exercício: {e}")
        time.sleep(2)

    while linha <= total_rows:
        select_element = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable(
                (By.XPATH, "/html/body/div/div/div[3]/div[1]/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td/form/table[1]/tbody/tr[1]/td[2]/select"))
            )
        select = Select(select_element)
        
        try:
            select.select_by_visible_text('CNPJ')
        except Exception as e:
            print(f"Erro ao selecionar o exercício: {e}")
            time.sleep(2)
        
        Num_CDA = planilha.cell(row=linha, column=3).value
        Num_CNPJ = planilha.cell(row=linha, column=2).value
        
        if Num_CNPJ is not None:
            Num_CNPJ = re.sub(r'\D', '', str(Num_CNPJ))  # Remove non-digit characters
        
        if Num_CDA is None:
            print("Num_CDA está vazio. Saindo do loop.")
            break
        
        Campo_CNPJ = driver.find_element(By.XPATH, "/html/body/div/div/div[3]/div[1]/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td/form/table[1]/tbody/tr[2]/td[2]/input")
        Campo_CNPJ.clear()
        Campo_CNPJ.send_keys(Num_CNPJ)
        
        Campo_CDA = driver.find_element(By.XPATH, "//*[@id='id_numero_daf']")
        Campo_CDA.clear()
        Campo_CDA.send_keys(Num_CDA)
        
        btn_Consultar = driver.find_element(By.XPATH, "/html/body/div/div/div[3]/div[1]/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td/form/table[2]/tbody/tr/td/input[1]")
        btn_Consultar.click()
        time.sleep(1)
        
        try:
            Valor_Total_Elemento = driver.find_element(By.XPATH, "/html/body/div/div/div[3]/div[1]/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr[9]/td[2]/b/font")
            Valor_Total = Valor_Total_Elemento.text
            planilha.cell(row=linha, column=4, value=Valor_Total)
            workbook.save(excel_file)

            # Gera e salva o PDF dinamicamente com base no valor de Num_CDA
            gerar_pdf_dinamico(driver, caminho_diretorio, str(Num_CDA))

            linha += 1
            time.sleep(1.5)
            driver.back()  # Volta para a página anterior
            time.sleep(1)
        except Exception as e:
            print(f"Erro durante o processamento da linha {linha}: {e}")
            Mensagem_Erro = "Mensagem Broker: QUITADO NAO PERMITE PAGAMENTO. DUVIDA, CONTATE ADM.FAZENDARIA"
            planilha.cell(row=linha, column=4, value=Mensagem_Erro)
            workbook.save(excel_file)
            linha += 1
            time.sleep(1.5)
            driver.back()  # Volta para a página anterior
            time.sleep(1)

    driver.quit()

# Função para abrir uma janela de seleção de arquivos e obter o caminho do arquivo Excel
def browse_file():
    filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls")])
    excel_file_entry.delete(0, tk.END)
    excel_file_entry.insert(0, filename)

# Função para abrir uma janela de seleção de diretório e obter o caminho do diretório onde salvar os PDFs
def browse_directory():
    directory = filedialog.askdirectory()
    directory_entry.delete(0, tk.END)
    directory_entry.insert(0, directory)

# Função que inicia o processo em uma nova thread
def on_start():
    excel_file = excel_file_entry.get()
    sheet_name = sheet_name_entry.get()
    start_row = start_row_entry.get()
    caminho_diretorio = directory_entry.get()

    # Verifica se todos os campos estão preenchidos
    if not excel_file or not sheet_name or not start_row or not caminho_diretorio:
        messagebox.showwarning("Erro de Entrada", "Por favor, preencha todos os campos.")
        return

    # Inicia a função start_process em uma nova thread para não bloquear a interface gráfica
    threading.Thread(target=start_process, args=(excel_file, sheet_name, start_row, caminho_diretorio)).start()

# Criação da janela GUI com tkinter
root = tk.Tk()
root.title("Automação de CDA PGE-MG")

# Adiciona os elementos da interface gráfica (labels, campos de entrada, botões)
tk.Label(root, text="Arquivo Excel:").grid(row=0, column=0, padx=10, pady=5, sticky=tk.W)
excel_file_entry = tk.Entry(root, width=100)
excel_file_entry.grid(row=0, column=1, padx=10, pady=5)
tk.Button(root, text="Buscar", command=browse_file).grid(row=0, column=2, padx=10, pady=5)

tk.Label(root, text="Nome da Planilha:").grid(row=1, column=0, padx=10, pady=5, sticky=tk.W)
sheet_name_entry = tk.Entry(root, width=100)
sheet_name_entry.grid(row=1, column=1, padx=10, pady=5)

tk.Label(root, text="Linha Inicial:").grid(row=2, column=0, padx=10, pady=5, sticky=tk.W)
start_row_entry = tk.Entry(root, width=100)
start_row_entry.grid(row=2, column=1, padx=10, pady=5)

tk.Label(root, text="Diretório para Salvar PDFs:").grid(row=3, column=0, padx=10, pady=5, sticky=tk.W)
directory_entry = tk.Entry(root, width=100)
directory_entry.grid(row=3, column=1, padx=10, pady=5)
tk.Button(root, text="Buscar Diretório", command=browse_directory).grid(row=3, column=2, padx=10, pady=5)

# Botão para iniciar o processo
tk.Button(root, text="Iniciar Processo", command=on_start).grid(row=4, column=0, columnspan=3, pady=10)

# Inicia o loop principal da interface gráfica
root.mainloop()
